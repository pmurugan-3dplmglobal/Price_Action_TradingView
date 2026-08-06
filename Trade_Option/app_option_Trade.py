import os, json, csv, time, threading, subprocess, sys, signal, logging
COMMON_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "common"))
if COMMON_DIR not in sys.path:
    sys.path.insert(0, COMMON_DIR)
from datetime import datetime as dt, time as datetime_time
from flask import Flask, render_template_string, jsonify, request, Response
import trade_db
from trading_core import (
    lookup_scan_sl_target,
    derive_sl_targets_for_contract,
    close_position as shared_close_position,
    close_stock_position as shared_close_stock_position,
    clear_executed_exit,
    log_to_journal
)

app = Flask(__name__)

# ──────────────────────────────────────────────
#  FILE PATHS & DASHBOARD CONFIG
# ──────────────────────────────────────────────


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "input", "program_config.json")

STATE_FILE = os.path.join(BASE_DIR, "output", "monitor", "stock_positions_state.json")
JOURNAL_FILE = os.path.join(BASE_DIR, "output", "monitor", "trade_journal.csv")
INDEX_LOG_FILE = os.path.join(BASE_DIR, "output", "logs", "bull_index_trade_engine.log")
NIFTY50_LOG_FILE = os.path.join(BASE_DIR, "output", "logs", "bull_nifty50_scanner.log")
DAILY_LOG_FILE = os.path.join(BASE_DIR, "output", "logs", "bull_daily_scanner.log")
SCAN_DISPLAY_FILE = os.path.join(BASE_DIR, "output", "monitor", "scan_display_data.json")
SCAN_DISPLAY_INDEX_FILE = os.path.join(BASE_DIR, "output", "monitor", "scan_display_index.json")
LIVE_EXECUTION_FLAG = os.path.join(BASE_DIR, "input", "nifty50_live.flag")
LIVE_EXECUTION_FLAG_INDEX = os.path.join(BASE_DIR, "input", "index_live.flag")

DASHBOARD_PORT = 6060
REFRESH_SECONDS = 1
ACTIVE_EDIT_LOCKS = set()

PROGRAMS = {
    "index": {
        "name": "Index Options Trade Engine",
        "file": "index_options_trade_engine.py",
        "desc": "Real-time index options intraday trading (NIFTY, BANKNIFTY, SENSEX)",
        "color": "#58a6ff",
        "log_file": INDEX_LOG_FILE,
        "config_fields": {
            "timeframe_entry": {"label": "Entry Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "3minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "15minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 15},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0},
            "strike_range": {"label": "Strike Range (±)", "type": "number", "default": 0}
        }
    },
    "nifty50": {
        "name": "Stock Options Trade Engine",
        "file": "stock_options_trade_engine.py",
        "desc": "Scans Nifty 50 stock options, picks best setup & executes",
        "color": "#3fb950",
        "log_file": NIFTY50_LOG_FILE,
        "config_fields": {
            "timeframe_entry": {"label": "Entry Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "15minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "30minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 300},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0},
            "strike_range": {"label": "Strike Range (±)", "type": "number", "default": 0}
        }
    }
}

processes = {}
process_lock = threading.Lock()

data_lock = threading.Lock()
cached_data = {
    "positions": {},
    "journal": [],
    "log_tail": {pid: [] for pid in PROGRAMS},
    "stats": {"total_trades": 0, "win_rate": 0, "active_positions": 0, "pnl": 0},
    "scans": {"index": [], "nifty50": []},
    "scan_summary": {"index": {"anchors": {}, "abc_matches": {}}, "nifty50": {"anchors": {}, "abc_matches": {}}},
    "all_trades": [],
    "active_positions": [],
    "ltp": {},
    "anchor_status": {"running": False, "engine": None, "requested_at": None, "completed_at": None},
    "scan_display": {"date": "", "timestamp": "", "staged_trades": [], "active_positions": []},
    "live_execution": False,
    "live_execution_index": False
}
_ltp_last_fetch = 0
_last_scan_reset = ""

# ──────────────────────────────────────────────
#  PROCESS MANAGEMENT (Start/Stop Programs)
# ──────────────────────────────────────────────

def get_pid_for_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            return p.pid
    return None

def start_program(prog_id):
    token = check_token_valid()
    if not token["valid"]:
        print(f"Cannot start {prog_id}: {token['reason']}")
        return False
    with process_lock:
        if prog_id in processes and processes[prog_id].poll() is None:
            return False
        script_path = os.path.join(BASE_DIR, PROGRAMS[prog_id]["file"])
        try:
            p = subprocess.Popen(
                [sys.executable, script_path],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                cwd=BASE_DIR, creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            )
            processes[prog_id] = p
            return True
        except Exception as e:
            print(f"Failed to start {prog_id}: {e}")
            return False

def stop_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            if os.name == "nt":
                subprocess.run(["taskkill", "/F", "/T", "/PID", str(p.pid)], capture_output=True)
            else:
                os.kill(p.pid, signal.SIGTERM)
            processes.pop(prog_id, None)
            return True
    return False

# ──────────────────────────────────────────────
#  CONFIGURATION (program_config.json)
# ──────────────────────────────────────────────

def load_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE) as f:
                return json.load(f)
    except:
        pass
    return {}

def save_config(prog_id, data):
    cfg = load_config()
    cleaned = {}
    for k, v in data.items():
        if isinstance(v, str):
            try:
                if "." in v:
                    cleaned[k] = float(v)
                else:
                    cleaned[k] = int(v)
            except (ValueError, TypeError):
                cleaned[k] = v
        else:
            cleaned[k] = v
    cfg[prog_id] = cleaned
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return True

def get_backtest_mode():
    return load_config().get("_backtest", False)

def set_backtest_mode(enabled):
    cfg = load_config()
    cfg["_backtest"] = enabled
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# ──────────────────────────────────────────────
#  DATA LOADING & VALIDATION
# ──────────────────────────────────────────────

def check_token_valid():
    return {"valid": True, "reason": "Free Open-Source Data Feed Active"}

# ──────────────────────────────────────────────
#  DATA LOADING (trade_db, journal, logs)
# ──────────────────────────────────────────────

def load_positions():
    try:
        active = trade_db.get_active_trades()
        return {t["symbol"]: t for t in active}
    except:
        return {}

def load_journal():
    rows = []
    if os.path.exists(JOURNAL_FILE):
        try:
            with open(JOURNAL_FILE, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter="\t")
                seen = set()
                for row in reader:
                    key = (row.get("Symbol", ""), row.get("Timestamp", ""))
                    if key not in seen:
                        seen.add(key)
                        rows.append(row)
        except Exception:
            pass
    return rows[-200:]

def get_best_log_file(filepath):
    if not filepath:
        return ""
    candidates = [
        os.path.join(BASE_DIR, "output", "logs", os.path.basename(filepath)),
        os.path.join(BASE_DIR, filepath),
        filepath,
    ]
    existing = []
    for c in candidates:
        if os.path.exists(c):
            existing.append((os.path.getmtime(c), c))
    if existing:
        existing.sort(key=lambda x: x[0], reverse=True)
        return existing[0][1]
    return filepath

def tail_log(filepath, n=200):
    best_file = get_best_log_file(filepath)
    if not best_file or not os.path.exists(best_file):
        return []
    try:
        with open(best_file, encoding="utf-8") as f:
            lines = f.readlines()
        return lines[-n:]
    except Exception:
        return []

def compute_stats(positions, journal):
    active = len(positions)
    total = len(journal)
    wins = sum(1 for j in (journal or []) if str(j.get("P&L %") or "").replace("%", "").replace("-", "").strip()
               and str(j.get("Action") or "").startswith("EXIT_"))
    win_rate = round((wins / total) * 100, 1) if total > 0 else 0
    pnl = 0.0
    for j in (journal or []):
        try:
            pnl_str = str(j.get("P&L %") or "").replace("%", "")
            if pnl_str and pnl_str != "-":
                pnl += float(pnl_str)
        except Exception:
            pass
    return {"total_trades": total, "win_rate": win_rate, "active_positions": active, "pnl": round(pnl, 2)}

SCAN_SYMBOLS = [
    "RELIANCE","TCS","HDFCBANK","ICICIBANK","INFY","ITC","SBIN","BHARTIARTL","LT","WIPRO",
    "ADANIENT","ADANIPORTS","APOLLOHOSP","ASIANPAINT","AXISBANK","BAJAJ-AUTO","BAJAJFINSV",
    "BAJFINANCE","BEL","CIPLA","COALINDIA","DRREDDY","EICHERMOT","ETERNAL","GRASIM","HCLTECH",
    "HDFCLIFE","HINDALCO","HINDUNILVR","INDIGO","JIOFIN","JSWSTEEL",
    "KOTAKBANK","M&M","MARUTI","MAXHEALTH","NESTLEIND","NTPC","ONGC","POWERGRID","SBILIFE",
    "SHRIRAMFIN","SUNPHARMA","TATACONSUM","TMPV","TATASTEEL","TECHM","TITAN","TRENT","ULTRACEMCO",
    "NIFTY","BANKNIFTY"
]

# ──────────────────────────────────────────────
#  SCAN PARSING — Split anchor vs ABC per symbol
# ──────────────────────────────────────────────

def _extract_symbol(line):
    for sym in SCAN_SYMBOLS:
        if sym in line:
            return sym
    return None

def parse_scans_for_program(log_lines, prog_id):
    matches = []
    anchors = {}
    abc_matches = {}
    for line in log_lines:
        clean = line.strip()
        if "ANCHOR" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                anchors[sym] = clean
        elif "MATCH" in clean or "BEST TRADE" in clean or "Match" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                abc_matches[sym] = clean
    return matches, anchors, abc_matches

# ──────────────────────────────────────────────
#  BACKGROUND DATA REFRESH THREAD
# ──────────────────────────────────────────────

def refresh_data():
    global cached_data, _ltp_last_fetch, _last_scan_reset
    while True:
        with data_lock:
            pos = load_positions()
            journal = load_journal()
            cached_data["positions"] = pos
            cached_data["journal"] = journal
            cached_data["stats"] = compute_stats(pos, journal)
            try:
                all_t = trade_db.get_all_trades()
                cached_data["all_trades"] = all_t
            except Exception:
                cached_data["all_trades"] = []
            for pid in PROGRAMS:
                log_file = PROGRAMS[pid].get("log_file")
                log_lines = tail_log(log_file) if log_file else []
                cached_data["log_tail"][pid] = log_lines
                scan_lines, anchors, abc_matches = parse_scans_for_program(log_lines, pid)
                cached_data["scans"][pid] = scan_lines
                cached_data["scan_summary"][pid] = {"anchors": anchors, "abc_matches": abc_matches}
            now_ist = dt.now()
            today_str = now_ist.strftime("%Y-%m-%d")
            market_open = now_ist.replace(hour=9, minute=0, second=0, microsecond=0)
            if _last_scan_reset != today_str and now_ist >= market_open:
                _last_scan_reset = today_str
                for f in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
                    try:
                        if os.path.exists(f):
                            with open(f, "r") as fh:
                                existing = json.load(fh)
                            if existing and isinstance(existing, dict):
                                existing["date"] = today_str
                                existing["timestamp"] = now_ist.strftime("%Y-%m-%d %H:%M:%S")
                                if "staged_trades" not in existing: existing["staged_trades"] = []
                                if "carry_forward" not in existing: existing["carry_forward"] = []
                                with open(f, "w") as fh:
                                    json.dump(existing, fh, indent=2)
                                continue  # Keep existing scan display data intact
                        empty_scan = {"date": today_str, "timestamp": now_ist.strftime("%Y-%m-%d %H:%M:%S"), "staged_trades": [], "carry_forward": [], "active_live": []}
                        with open(f, "w") as fh:
                            json.dump(empty_scan, fh)
                    except Exception:
                        pass
            scan_display = {}
            try:
                if os.path.exists(SCAN_DISPLAY_FILE):
                    with open(SCAN_DISPLAY_FILE, "r") as f:
                        scan_display["nifty50"] = json.load(f)
                stock_scan_file = os.path.abspath(os.path.join(BASE_DIR, "..", "Trade_Stock", "output", "monitor", "scan_display_data.json"))
                if os.path.exists(stock_scan_file):
                    with open(stock_scan_file, "r") as f:
                        stock_disp = json.load(f)
                    if stock_disp.get("staged_trades"):
                        if not scan_display.get("nifty50"):
                            scan_display["nifty50"] = stock_disp
                        else:
                            existing_staged = scan_display["nifty50"].get("staged_trades", [])
                            existing_syms = {t.get("symbol") for t in existing_staged if t.get("symbol")}
                            for st in stock_disp.get("staged_trades", []):
                                if st.get("symbol") not in existing_syms:
                                    existing_staged.append(st)
                            scan_display["nifty50"]["staged_trades"] = existing_staged
            except Exception:
                pass
            try:
                if os.path.exists(SCAN_DISPLAY_INDEX_FILE):
                    with open(SCAN_DISPLAY_INDEX_FILE, "r") as f:
                        scan_display["index"] = json.load(f)
            except Exception:
                pass
            cached_data["scan_display"] = scan_display
            cached_data["live_execution"] = os.path.exists(LIVE_EXECUTION_FLAG)
            cached_data["live_execution_index"] = os.path.exists(LIVE_EXECUTION_FLAG_INDEX)
            try:
                active = trade_db.get_active_trades()
                active_list = []
                for t in active:
                    active_list.append({
                        "contract": t.get("contract") or t.get("symbol"),
                        "symbol": t.get("symbol") or t.get("contract"),
                        "quantity": t.get("position_size", 1),
                        "entry_price": t.get("entry_spot", 0),
                        "entry_spot": t.get("entry_spot", 0),
                        "ltp": t.get("current_price") or t.get("entry_spot", 0),
                        "pnl": t.get("pnl", 0),
                        "current_sl": t.get("current_sl", 0),
                        "t1": t.get("t1", 0),
                        "t2": t.get("t2", 0),
                        "t3": t.get("t3", 0),
                        "pattern": t.get("pattern", "ACTIVE"),
                        "source": "local"
                    })
                cached_data["active_positions"] = active_list
            except Exception:
                pass
        if int(time.time()) % 3600 < REFRESH_SECONDS:
            auto_export_if_new_month()
        time.sleep(REFRESH_SECONDS)

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Price Action Option Strategy — TradingView Edition</title>
    <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
    <script>
        // ── TradingView Chart Helper ──
        function openTVChart(symbol) {
            const cleanSym = (symbol || 'NIFTY').replace(/\\s+/g, '').replace('^', '').toUpperCase();
            let tvSymbol = "NSE:" + cleanSym;
            if (cleanSym === 'NIFTY') tvSymbol = "NSE:NIFTY";
            else if (cleanSym === 'BANKNIFTY') tvSymbol = "NSE:BANKNIFTY";
            else if (cleanSym === 'SENSEX') tvSymbol = "BSE:SENSEX";

            const inp = document.getElementById('tv-symbol-input');
            if (inp) inp.value = cleanSym;

            switchLeftTab('tvchart-tab');

            const container = document.getElementById('tv_chart_container');
            if (container) {
                container.innerHTML = '';
                new TradingView.widget({
                    "autosize": true,
                    "symbol": tvSymbol,
                    "interval": "15",
                    "timezone": "Asia/Kolkata",
                    "theme": "dark",
                    "style": "1",
                    "locale": "en",
                    "toolbar_bg": "#f1f3f6",
                    "enable_publishing": false,
                    "allow_symbol_change": true,
                    "container_id": "tv_chart_container"
                });
            }
        }
        function loadTVChartFromInput() {
            const sym = (document.getElementById('tv-symbol-input') || {}).value || 'NIFTY';
            openTVChart(sym);
        }

        // ── Filter State ──
        let journalFilter = 'all';
        let scanFilter = 'all';
        let logFilter = 'all';
        let positionFilter = 'active';

        // ── Tab Switching ──
        function switchLeftTab(tabId, btnEl) {
            document.querySelectorAll('.left-tab-content').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.left-tab-btn').forEach(b => b.classList.remove('active'));
            const content = document.getElementById(tabId);
            if (content) content.classList.add('active');
            const b = btnEl || (window.event ? (window.event.currentTarget || window.event.target) : null);
            if (b && b.classList) b.classList.add('active');
            if (tabId === 'backtest-tab') renderBacktest();
            if (tabId === 'scan-tab-left') renderScanTab();
        }

        function switchTab(tabId) {
            document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            event.target.classList.add('active');
        }

        // ── Filter Controls ──
        function setFilter(type, value) {
            if (type === 'journal') journalFilter = value;
            else if (type === 'scan') scanFilter = value;
            else if (type === 'log') logFilter = value;
            else if (type === 'position') { positionFilter = value; document.querySelectorAll('.pos-filter-btn').forEach(b => b.classList.remove('active')); event.target.classList.add('active'); }
            renderReport();
        }

        // ── Backtest Controls ──
        async function toggleBacktestMode(enabled) {
            try {
                await fetch('/api/backtest/mode', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({enabled: enabled})
                });
                refreshBacktestMode();
            } catch(e) { console.log(e); }
        }

        // ── Live Execution Toggle ──
        async function toggleLiveExecution(engine) {
            const isIndex = engine === 'index';
            const tgl = document.getElementById(isIndex ? 'live-exec-toggle-idx' : 'live-exec-toggle');
            const lbl = document.getElementById(isIndex ? 'live-exec-label-idx' : 'live-exec-label');
            const was = tgl.classList.contains('active');
            const ep = isIndex ? '/api/live-execution/index' : '/api/live-execution/nifty50';
            try {
                const r = await fetch(ep, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({enabled: !was})
                });
                const d = await r.json();
                if (d.ok) {
                    tgl.classList.toggle('active', d.enabled);
                    const label = isIndex ? 'IDX' : 'N50';
                    lbl.textContent = d.enabled ? label+': ON' : label+': OFF';
                    lbl.style.color = d.enabled ? '#3fb950' : '#8b949e';
                }
            } catch(e) { console.log(e); }
        }

        async function toggleCardLive(pid) {
            const ep = '/api/live-execution/' + pid;
            const sw = document.getElementById('live-toggle-' + pid);
            const lb = document.getElementById('live-label-' + pid);
            const was = sw.classList.contains('on');
            try {
                const r = await fetch(ep, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({enabled: !was})
                });
                const d = await r.json();
                if (d.ok) {
                    sw.classList.toggle('on', d.enabled);
                    lb.textContent = d.enabled ? 'LIVE' : 'SCAN';
                    lb.style.color = d.enabled ? '#3fb950' : '#8b949e';
                }
            } catch(e) { console.log(e); }
        }

        function refreshLiveExec() {
            const d = window._lastData;
            if (!d) return;
            const enabled = d.live_execution || false;
            const enabledIdx = d.live_execution_index || false;
            const tgl = document.getElementById('live-exec-toggle');
            const lbl = document.getElementById('live-exec-label');
            if (tgl) {
                tgl.classList.toggle('active', enabled);
                lbl.textContent = enabled ? 'N50: ON' : 'N50: OFF';
                lbl.style.color = enabled ? '#3fb950' : '#8b949e';
            }
            const tglIdx = document.getElementById('live-exec-toggle-idx');
            const lblIdx = document.getElementById('live-exec-label-idx');
            if (tglIdx) {
                tglIdx.classList.toggle('active', enabledIdx);
                lblIdx.textContent = enabledIdx ? 'IDX: ON' : 'IDX: OFF';
                lblIdx.style.color = enabledIdx ? '#3fb950' : '#8b949e';
            }
        }

        // ── Program Start/Stop ──
        async function toggleProgram(progId, action) {
            if (action === 'stop') {
                try { await fetch('/api/anchor/stop', {method: 'POST'}); } catch(e) {}
            }
            const btn = document.querySelector(`.start-btn[onclick*="'${progId}'"]`);
            const fb = document.getElementById(`cfg-fb-${progId}`);
            if (action === 'start' && btn) btn.disabled = true;
            try {
                const r = await fetch(`/api/programs/${progId}/${action}`, {method: 'POST'});
                const d = await r.json();
                if (d.ok) {
                    if (fb) { fb.textContent = action === 'start' ? 'Running!' : 'Stopped'; fb.style.color = '#3fb950'; setTimeout(() => { fb.textContent = ''; }, 2000); }
                    setTimeout(refreshData, 500);
                }
                if (d.error) {
                    if (fb) { fb.textContent = d.error; fb.style.color = '#f85149'; setTimeout(() => { fb.textContent = ''; }, 4000); }
                    if (action === 'start' && btn) btn.disabled = false;
                }
            } catch(e) { console.log(e); if (action === 'start' && btn) btn.disabled = false; }
        }

        function toggleConfig(headerEl) {
            const body = headerEl.parentElement.querySelector('.config-body');
            const arrow = headerEl.querySelector('.config-arrow');
            if (body.style.display === 'block') {
                body.style.display = 'none';
                arrow.textContent = '\u25B6';
            } else {
                body.style.display = 'block';
                arrow.textContent = '\u25BC';
            }
        }

        async function saveConfig(progId) {
            const inputs = document.querySelectorAll(`.config-input[data-prog="${progId}"]`);
            const data = {};
            inputs.forEach(inp => {
                data[inp.getAttribute('data-field')] = inp.value;
            });
            try {
                const r = await fetch(`/api/config/${progId}`, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                const d = await r.json();
                const fb = document.getElementById(`cfg-fb-${progId}`);
                if (d.ok) {
                    fb.textContent = 'Saved!';
                    fb.style.color = '#3fb950';
                } else {
                    fb.textContent = 'Failed';
                    fb.style.color = '#f85149';
                }
                setTimeout(() => { fb.textContent = ''; }, 2000);
            } catch(e) { console.log(e); }
        }

        async function clearScanData() {
            await fetch('/api/scan/clear', {method:'POST'});
            await refreshData();
            renderScanTab();
        }

        async function scanExport() {
            try {
                const r = await fetch('/api/scan/export', {method:'POST'});
                if (!r.ok) return;
                const blob = await r.blob();
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                const now = new Date();
                const pad = n => String(n).padStart(2, '0');
                const dd = pad(now.getDate());
                const mm = pad(now.getMonth() + 1);
                const yy = String(now.getFullYear()).slice(-2);
                const hhmin = pad(now.getHours()) + pad(now.getMinutes());
                a.download = `scan_export_${dd}_${mm}_${yy}_${hhmin}.csv`;
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                URL.revokeObjectURL(url);
            } catch(e) { console.log(e); }
        }

        let _sortPref = {};

        function sortTable(th, colIdx) {
            const table = th.closest('table');
            const container = table.closest('[id]');
            const key = container ? container.id : '_default';
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const dir = th.getAttribute('data-dir') === 'asc' ? 'desc' : 'asc';
            th.setAttribute('data-dir', dir);
            _sortPref[key] = { col: colIdx, dir: dir };
            rows.sort((a, b) => {
                let aVal = a.cells[colIdx].innerText.trim();
                let bVal = b.cells[colIdx].innerText.trim();
                let aNum = parseFloat(aVal.replace('%', '').replace(',', ''));
                let bNum = parseFloat(bVal.replace('%', '').replace(',', ''));
                if (!isNaN(aNum) && !isNaN(bNum)) {
                    return dir === 'asc' ? aNum - bNum : bNum - aNum;
                }
                return dir === 'asc' ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
            });
            rows.forEach(r => tbody.appendChild(r));
        }

        function reapplySorts() {
            for (const [key, pref] of Object.entries(_sortPref)) {
                const container = document.getElementById(key);
                if (!container) continue;
                container.querySelectorAll('table').forEach(table => {
                    const ths = table.querySelectorAll('th');
                    if (pref.col >= ths.length) return;
                    const th = ths[pref.col];
                    th.setAttribute('data-dir', pref.dir);
                    const tbody = table.querySelector('tbody');
                    if (!tbody) return;
                    const rows = Array.from(tbody.querySelectorAll('tr'));
                    rows.sort((a, b) => {
                        let aVal = a.cells[pref.col].innerText.trim();
                        let bVal = b.cells[pref.col].innerText.trim();
                        let aNum = parseFloat(aVal.replace('%', '').replace(',', ''));
                        let bNum = parseFloat(bVal.replace('%', '').replace(',', ''));
                        if (!isNaN(aNum) && !isNaN(bNum)) {
                            return pref.dir === 'asc' ? aNum - bNum : bNum - aNum;
                        }
                        return pref.dir === 'asc' ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
                    });
                    rows.forEach(r => tbody.appendChild(r));
                });
            }
        }

        async function buyScannedTrade(btnEl, symbol, contract, side, entry, sl, t1, t2, t3, engine) {
            const dispName = contract || symbol;
            if (!confirm(`Confirm 1-Click BUY for ${dispName}?`)) return;
            if (btnEl) {
                btnEl.disabled = true;
                btnEl.textContent = 'BUYING...';
                btnEl.style.background = '#8b949e';
            }
            try {
                const r = await fetch('/api/buy-scanned-trade', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({symbol, contract, side, entry_spot: entry, current_sl: sl, t1, t2, t3, engine})
                });
                const d = await r.json();
                if (d.ok) {
                    showToast(d.message || `BUY order placed for ${dispName}`, 'success');
                    setTimeout(refreshData, 500);
                } else {
                    showToast(`Failed: ${d.error || 'unknown error'}`, 'error');
                    if (btnEl) {
                        btnEl.disabled = false;
                        btnEl.textContent = 'BUY';
                        btnEl.style.background = '#2ea043';
                    }
                }
            } catch(e) {
                showToast(`Network error: ${e.message}`, 'error');
                if (btnEl) {
                    btnEl.disabled = false;
                    btnEl.textContent = 'BUY';
                    btnEl.style.background = '#2ea043';
                }
            }
        }

        window.manualExitPosition = async function(btnEl, contract, engine) {
            if (!confirm(`Confirm Manual EXIT for ${contract}?`)) return;
            if (btnEl) {
                btnEl.disabled = true;
                btnEl.textContent = 'EXITING...';
                btnEl.style.background = '#8b949e';
            }
            try {
                const r = await fetch('/api/exit-position', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({contract, symbol: contract, engine})
                });
                const d = await r.json();
                if (d.ok) {
                    showToast(d.message || `Position EXITED for ${contract}`, 'success');
                    setTimeout(refreshData, 500);
                } else {
                    showToast(`Exit failed: ${d.error || 'unknown error'}`, 'error');
                    if (btnEl) {
                        btnEl.disabled = false;
                        btnEl.textContent = 'EXIT';
                        btnEl.style.background = '#da3633';
                    }
                }
            } catch(e) {
                showToast(`Network error: ${e.message}`, 'error');
                if (btnEl) {
                    btnEl.disabled = false;
                    btnEl.textContent = 'EXIT';
                    btnEl.style.background = '#da3633';
                }
            }
        };

        window.manualExitAllPositions = async function() {
            if (!confirm('Are you sure you want to EXIT ALL active positions?')) return;
            try {
                const r = await fetch('/api/exit-all-positions', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'}
                });
                const d = await r.json();
                if (d.ok) {
                    showToast(d.message || 'All active positions EXITED', 'success');
                    setTimeout(refreshData, 500);
                } else {
                    showToast(`Exit all failed: ${d.error || 'unknown error'}`, 'error');
                }
            } catch(e) {
                showToast(`Network error: ${e.message}`, 'error');
            }
        };

        async function runNegationAnalysis() {
            const symbol = (document.getElementById('an-symbol') || {}).value || '';
            const entryVal = (document.getElementById('an-entry') || {}).value;
            const entry = entryVal ? parseFloat(entryVal) : 0;
            const tf = (document.getElementById('an-tf') || {}).value || '75min';
            const eng = (document.getElementById('an-engine') || {}).value || 'nifty50';
            const resBox = document.getElementById('analyzer-results');
            const btn = document.getElementById('an-submit-btn');

            if (!symbol) {
                showToast('Please enter a valid Symbol or Contract name (e.g. WIPRO26AUG200CE or VEDL)', 'error');
                return;
            }

            if (btn) {
                btn.disabled = true;
                btn.textContent = 'Analyzing Negation Targets...';
            }
            if (resBox) {
                resBox.innerHTML = '<div style="padding:20px;text-align:center;color:#8b949e;">Fetching chart candles and computing Negation Theory pivots...</div>';
            }

            try {
                const r = await fetch('/api/analyze-trade', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({symbol: symbol, entry_price: entry, timeframe: tf, engine: eng})
                });
                const d = await r.json();
                if (btn) {
                    btn.disabled = false;
                    btn.textContent = '🔍 Analyze Negation Targets & SL';
                }

                if (d.ok) {
                    window._lastAnalysis = d;
                    const t1 = d.t1 !== 'N/A' && d.t1 ? parseFloat(d.t1).toFixed(2) : 'N/A';
                    const t2 = d.t2 !== 'N/A' && d.t2 ? parseFloat(d.t2).toFixed(2) : 'N/A';
                    const t3 = d.t3 !== 'N/A' && d.t3 ? parseFloat(d.t3).toFixed(2) : 'N/A';
                    const sl = d.current_sl ? parseFloat(d.current_sl).toFixed(2) : 'N/A';

                    if (resBox) {
                        resBox.innerHTML = `
                            <div style="background:#161b22;border:1px solid #30363d;border-radius:8px;padding:16px;margin-top:16px;">
                                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                                    <h3 style="color:#58a6ff;margin:0;font-size:15px;">${d.symbol} — Negation Theory Analysis</h3>
                                    <span class="badge badge-open">${d.pattern}</span>
                                </div>
                                <div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(130px, 1fr));gap:12px;margin-bottom:16px;">
                                    <div style="background:#0d1117;padding:10px;border-radius:6px;border:1px solid #30363d;">
                                        <div style="font-size:11px;color:#8b949e;">Entry Price</div>
                                        <div style="font-size:16px;font-weight:bold;color:#c9d1d9;">₹${d.entry_price.toFixed(2)}</div>
                                    </div>
                                    <div style="background:#0d1117;padding:10px;border-radius:6px;border:1px solid #f8514944;">
                                        <div style="font-size:11px;color:#f85149;">Active SL (10% Rule)</div>
                                        <div style="font-size:16px;font-weight:bold;color:#f85149;">₹${sl}</div>
                                    </div>
                                    <div style="background:#0d1117;padding:10px;border-radius:6px;border:1px solid #3fb95044;">
                                        <div style="font-size:11px;color:#3fb950;">Target 1 (T1)</div>
                                        <div style="font-size:16px;font-weight:bold;color:#3fb950;">${t1 !== 'N/A' ? '₹'+t1 : 'N/A'}</div>
                                    </div>
                                    <div style="background:#0d1117;padding:10px;border-radius:6px;border:1px solid #3fb95044;">
                                        <div style="font-size:11px;color:#3fb950;">Target 2 (T2)</div>
                                        <div style="font-size:16px;font-weight:bold;color:#3fb950;">${t2 !== 'N/A' ? '₹'+t2 : 'N/A'}</div>
                                    </div>
                                    <div style="background:#0d1117;padding:10px;border-radius:6px;border:1px solid #3fb95044;">
                                        <div style="font-size:11px;color:#3fb950;">Target 3 (T3)</div>
                                        <div style="font-size:16px;font-weight:bold;color:#3fb950;">${t3 !== 'N/A' ? '₹'+t3 : 'N/A'}</div>
                                    </div>
                                    <div style="background:#0d1117;padding:10px;border-radius:6px;border:1px solid #58a6ff44;">
                                        <div style="font-size:11px;color:#58a6ff;">Risk-Reward (RR)</div>
                                        <div style="font-size:16px;font-weight:bold;color:#58a6ff;">${d.rr} : 1</div>
                                    </div>
                                </div>
                                <div style="display:flex;gap:10px;">
                                    <button onclick="applyAnalyzedTradeToActive()" style="background:#2ea043;color:#ffffff;border:none;padding:8px 16px;border-radius:6px;font-weight:bold;cursor:pointer;font-size:12px;">📌 Apply to Active Positions</button>
                                </div>
                            </div>
                        `;
                    }
                } else {
                    if (resBox) resBox.innerHTML = `<div style="padding:16px;background:#f8514922;border:1px solid #f85149;border-radius:6px;color:#f85149;">Error: ${d.error}</div>`;
                }
            } catch(e) {
                if (btn) {
                    btn.disabled = false;
                    btn.textContent = '🔍 Analyze Negation Targets & SL';
                }
                if (resBox) resBox.innerHTML = `<div style="padding:16px;background:#f8514922;border:1px solid #f85149;border-radius:6px;color:#f85149;">Network Error: ${e.message}</div>`;
            }
        }

        async function applyAnalyzedTradeToActive() {
            const d = window._lastAnalysis;
            if (!d) return;
            try {
                const r = await fetch('/api/update-position', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        engine: d.engine,
                        symbol: d.symbol,
                        current_sl: d.current_sl,
                        t1: d.t1 !== 'N/A' ? d.t1 : null,
                        t2: d.t2 !== 'N/A' ? d.t2 : null,
                        t3: d.t3 !== 'N/A' ? d.t3 : null
                    })
                });
                const res = await r.json();
                if (res.ok) {
                    showToast(`Position ${d.symbol} added to Active Positions!`, 'success');
                    setTimeout(refreshData, 500);
                } else {
                    showToast(`Failed: ${res.error}`, 'error');
                }
            } catch(e) {
                showToast(`Error: ${e.message}`, 'error');
            }
        }

        let editStates = {};
        window._isEditing = false;
        function renderScanTab(force=false) {
            if (!force && (window._isEditing || (document.activeElement && document.activeElement.tagName === "INPUT"))) return;
            const d = window._lastData;
            if (!d) return;
            const sd = d.scan_display || {};
            const activeContracts = new Set();
            (d.active_positions || []).forEach(p => {
                const c = (p.contract || p.symbol || '').replace(/\\s+/g, '').toUpperCase();
                if (c) activeContracts.add(c);
            });
            (d.all_trades || []).forEach(t => {
                if ((t.status || '').toLowerCase() === 'active') {
                    const c = (t.contract || t.symbol || '').replace(/\\s+/g, '').toUpperCase();
                    if (c) activeContracts.add(c);
                }
            });

            const filter = (document.getElementById('scan-engine-filter') || {}).value || 'all';
            let scanHtml = '';
            const colHeaders = '<th onclick="sortTable(this,0)">Symbol</th><th onclick="sortTable(this,1)">Contract</th><th onclick="sortTable(this,2)">Side</th><th onclick="sortTable(this,3)">Entry</th><th onclick="sortTable(this,4)">SL</th><th onclick="sortTable(this,5)">T1</th><th onclick="sortTable(this,6)">T2</th><th onclick="sortTable(this,7)">T3</th><th onclick="sortTable(this,8)">AncherT</th><th onclick="sortTable(this,9)">EntryTime</th><th onclick="sortTable(this,10)">Result</th><th onclick="sortTable(this,11)">CF</th><th onclick="sortTable(this,12)">RR</th><th style="text-align:center">Action</th>';
            function tradeRow(t, resultBadge, eng) {
                const entry = t.entry_spot !== undefined && t.entry_spot !== null ? parseFloat(t.entry_spot).toFixed(2) : '-';
                const sl = t.current_sl !== undefined && t.current_sl !== null ? parseFloat(t.current_sl).toFixed(2) : '-';
                const t1v = t.t1 !== undefined && t.t1 !== null ? t.t1 : '-';
                const t2v = t.t2 !== undefined && t.t2 !== null ? t.t2 : '-';
                const t3v = t.t3 !== undefined && t.t3 !== null ? t.t3 : '-';
                const et = t.entry_time || '';
                let etFormatted = '-';
                if (et) {
                    const s = et.split('+')[0].replace('T', ' ');
                    const p = s.split(' ');
                    const dp = p[0] ? p[0].split('-') : [];
                    const tp = p[1] ? p[1].split(':') : [];
                    if (dp.length === 3 && tp.length >= 2)
                        etFormatted = `${dp[2]}-${dp[1]}-${dp[0].slice(-2)} ${tp[0]}:${tp[1]}`;
                }
                const at = t.candle_a_time || '';
                let atFormatted = '-';
                if (at) {
                    const s = at.split('+')[0].replace('T', ' ');
                    const p = s.split(' ');
                    const dp = p[0] ? p[0].split('-') : [];
                    const tp = p[1] ? p[1].split(':') : [];
                    if (dp.length === 3 && tp.length >= 2)
                        atFormatted = `${dp[2]}-${dp[1]}-${dp[0].slice(-2)} ${tp[0]}:${tp[1]}`;
                }
                let res = t.pattern || t.result || '-';
                if (res.includes('Engulf')) res = res.includes('Bear') || res.includes('BEAR') ? 'BEAR_ENG' : 'BULL_ENG';
                else if (res.includes('Two_Higher') || res.includes('Higher_Highs')) res = 'BULL_2HH';
                else if (res.includes('Two_Lower') || res.includes('Lower_Lows')) res = 'BEAR_2LL';
                else if (res.includes('HH_Sweep') || res.includes('HH_sweep')) res = 'BEAR_HH';
                else if (res.includes('Sweep') || res.includes('LL')) res = 'BULL_LL';
                else if (res.includes('Star') || res.includes('Shooting')) res = 'BEAR_STAR';
                else if (res.includes('Baby') || res.includes('Hammer')) res = 'BULL_HAM';
                else if (res.includes('Harami')) res = res.includes('Bear') || res.includes('BEAR') ? 'BEAR_HAR' : 'BULL_HAR';
                else if (res.includes('Base')) res = 'BULL_BASE';
                else if (res === 'SCAN_READY') res = 'BULL_ENG';
                const cf = t.carry_forward ? 'Yes' : 'No';
                const rr = t.rr !== undefined && t.rr !== null ? parseFloat(t.rr).toFixed(2) : '0.00';

                const symName = t.symbol || '';
                const symLink = `<a href="javascript:void(0)" onclick="openTVChart('${symName}')" style="color:#58a6ff;font-weight:bold;text-decoration:none;" title="Click to view TradingView chart">${symName}</a>`;
                let actCell = `<td style="text-align:center"><button class="btn-buy" onclick="openTVChart('${symName}')" style="background:#2962ff;color:#ffffff;border:none;padding:4px 12px;border-radius:4px;font-weight:bold;cursor:pointer;font-size:11px">CHART 📈</button></td>`;

                return `<tr><td>${symLink}</td><td style="font-size:11px">${t.contract||''}</td><td>${t.side||''}</td><td>${entry}</td><td>${sl}</td><td>${t1v}</td><td>${t2v}</td><td>${t3v}</td><td style="font-size:11px">${atFormatted}</td><td style="font-size:11px">${etFormatted}</td><td><span class="badge ${resultBadge}">${res}</span></td><td>${cf}</td><td>${rr}</td>${actCell}</tr>`;
            }
            const engines = filter === 'all' ? ['nifty50', 'index'] : [filter];
            engines.forEach(eng => {
                const data = sd[eng];
                if (!data) return;
                const rawStaged = (data.staged_trades || []).concat(data.carry_forward || []);
                const seenContracts = new Set();
                const staged = [];
                rawStaged.forEach(t => {
                    const key = (t.contract || t.symbol || '').trim();
                    if (key && !seenContracts.has(key)) {
                        seenContracts.add(key);
                        staged.push(t);
                    }
                });
                const engLabel = eng === 'nifty50' ? 'Nifty 50' : 'Index';
                if (staged.length) {
                    scanHtml += '<div class="scan-section-title">[' + engLabel + '] Scan Results (' + staged.length + ')</div>';
                    scanHtml += '<div style="overflow-x:auto"><table><thead><tr>' + colHeaders + '</tr></thead><tbody>';
                    staged.forEach(t => { scanHtml += tradeRow(t, 'badge-open', eng); });
                    scanHtml += '</tbody></table></div>';
                }
            });
            if (!scanHtml) {
                scanHtml = '<p class="empty-state">No scan trades yet. Run a scan cycle or wait for next cycle.</p>';
            } else {
                const ts = Object.values(sd).map(v => v.timestamp).filter(Boolean).sort().pop() || '-';
                scanHtml += '<div style="padding:6px 14px;font-size:11px;color:#8b949e">Last updated: ' + ts + '</div>';
            }
            document.getElementById('scan-body').innerHTML = scanHtml;
            reapplySorts();
        }

        function showToast(msg, type) {
            const c = document.getElementById('toast-container');
            if (!c) return;
            const t = document.createElement('div');
            t.className = 'toast toast-' + type;
            t.textContent = msg;
            c.appendChild(t);
            setTimeout(() => { t.style.opacity = '0'; setTimeout(() => t.remove(), 300); }, 2500);
        }
        function editRow(uid, symbol, sl, t1, t2, t3) {
            const clean = v => (v === '---' || v === '' || v === undefined || v === null) ? '' : v;
            editStates[uid] = {active: true, symbol: symbol, sl: clean(sl), t1: clean(t1), t2: clean(t2), t3: clean(t3)};
            window._isEditing = true;
            fetch('/api/edit-lock', {method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({symbol: symbol, active: true})});
            renderScanTab(true); renderReport(true);
        }
        function cancelEdit(uid) {
            const sym = editStates[uid]?.symbol;
            delete editStates[uid];
            if (Object.keys(editStates).length === 0) window._isEditing = false;
            fetch('/api/edit-lock', {method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({symbol: sym, active: false})});
            renderScanTab(true); renderReport(true);
        }
        async function saveEdit(uid, symbol, engine) {
            const es = editStates[uid];
            if (!es) return;
            const newSl = document.getElementById('sl_'+uid)?.value;
            const newT1 = document.getElementById('t1_'+uid)?.value;
            const newT2 = document.getElementById('t2_'+uid)?.value;
            const n            const actPosList = d.active_positions || d.positions || [];
            const journal = d.journal || [];

            const actPos = actPosList.length || 0;
            document.getElementById('stat-active').textContent = actPos;
            document.getElementById('stat-total').textContent = stats.total_trades || 0;

            let total = stats.total_trades || 0;
            let wins = 0;
            (journal || []).forEach(j => {
                const pnl = (j['P&L %'] || '').replace('%', '').replace('-', '').trim();
                if (pnl && (j.Action || '').startsWith('EXIT_')) wins++;
            });
            let wr = total > 0 ? ((wins/total)*100).toFixed(1) : 0;
            let wrEl = document.getElementById('stat-winrate');
            wrEl.textContent = wr + '%';
            wrEl.style.color = wr >= 50 ? '#3fb950' : '#f85149';

            let pnl = stats.pnl || 0;
            let pnlEl = document.getElementById('stat-pnl');
            pnlEl.textContent = pnl + '%';
            pnlEl.style.color = pnl >= 0 ? '#3fb950' : '#f85149';

            let posHtml = '';
            let allTrades = d.all_trades || [];
            let ltpData = d.ltp || {};
            let mergedPositions = [];
            let seenContracts = new Set();

            if (positionFilter === 'all' || positionFilter === 'active') {
                actPosList.forEach(kp => {
                    const c_name = kp.contract || kp.symbol;
                    if (!c_name || seenContracts.has(c_name)) return;
                    seenContracts.add(c_name);
                    const dbMatch = allTrades.find(t => {
                        const tc = (t.contract || '').replace(/\\s+/g, '').toUpperCase();
                        const ts = (t.symbol || '').replace(/\\s+/g, '').toUpperCase();
                        const kc = (c_name || '').replace(/\\s+/g, '').toUpperCase();
                        return (tc && (tc === kc || kc.includes(tc) || tc.includes(kc))) || (ts && (ts === kc || kc.includes(ts)));
                    });
                    const item = {
                        symbol: c_name,
                        contract: c_name,
                        engine: kp.exchange === 'NFO' ? 'Index' : 'Nifty 50',
                        pattern: kp.pattern || (dbMatch && dbMatch.pattern ? dbMatch.pattern : 'OPEN_TRADE'),
                        entry_spot: kp.entry_price,
                        quantity: kp.quantity,
                        pnl: kp.pnl,
                        current_sl: kp.current_sl !== undefined ? kp.current_sl : (dbMatch ? dbMatch.current_sl : ''),
                        t1: kp.t1 !== undefined ? kp.t1 : (dbMatch ? dbMatch.t1 : ''),
                        t2: kp.t2 !== undefined ? kp.t2 : (dbMatch ? dbMatch.t2 : ''),
                        t3: kp.t3 !== undefined ? kp.t3 : (dbMatch ? dbMatch.t3 : ''),
                        token: dbMatch ? (dbMatch.option_token || dbMatch.index_token || '') : (kp.token || ''),
                        status: 'ACTIVE',
                        source: 'local'
                    };
                    mergedPositions.push(item);
                });
            }

            const dbSeen = new Set();
            allTrades.forEach(t => {
                const contract = t.contract || t.symbol || '';
                const inActive = actPosList.some(kp => kp.contract === contract || kp.contract.includes(contract) || contract.includes(kp.contract));
                if (inActive) return;
                let st = (t.status || '').toLowerCase();
                if (st === 'active' && !inActive) st = 'exited';
                if (positionFilter === 'active' && st !== 'active') return;
                if (positionFilter === 'completed' && st !== 'sl_hit' && st !== 'target_hit' && st !== 'exited') return;
                if (positionFilter === 'sl_hit' && st !== 'sl_hit') return;

                const dedupKey = (contract + '_' + st + '_' + (t.entry_spot || '')).toUpperCase();
                if (dbSeen.has(dedupKey)) return;
                dbSeen.add(dedupKey);

                mergedPositions.push({
                    symbol: t.symbol || contract,
                    contract: contract,
                    engine: (t.engine === 'index' || (t.symbol||'').includes('NIFTY') || (t.symbol||'').includes('BANK')) ? 'Index' : 'Nifty 50',
                    pattern: t.pattern || '',
                    entry_spot: t.entry_spot !== undefined && t.entry_spot !== null ? t.entry_spot : '',
                    current_sl: t.current_sl !== undefined && t.current_sl !== null ? t.current_sl : '',
                    t1: t.t1 !== undefined && t.t1 !== null ? t.t1 : '',
                    t2: t.t2 !== undefined && t.t2 !== null ? t.t2 : '',
                    t3: t.t3 !== undefined && t.t3 !== null ? t.t3 : '',
                    status: st === 'exited' ? 'EXITED' : (t.status || 'ACTIVE'),
                    created_at: t.created_at || '',
                    exit_time: t.exit_time || '',
                    pnl_percent: t.pnl_percent,
                    token: t.option_token || t.index_token || '',
                    source: 'db'
                });
            });
            if (mergedPositions.length) {
                posHtml = '<table><thead><tr><th onclick="sortTable(this,0)">Symbol</th><th onclick="sortTable(this,1)">Source</th><th onclick="sortTable(this,2)">Pattern</th><th onclick="sortTable(this,3)">Entry</th><th onclick="sortTable(this,4)">SL</th><th onclick="sortTable(this,5)">T1</th><th onclick="sortTable(this,6)">T2</th><th onclick="sortTable(this,7)">T3</th><th onclick="sortTable(this,8)">LTP</th><th onclick="sortTable(this,9)">Qty</th><th onclick="sortTable(this,10)">Status</th><th onclick="sortTable(this,11)">P&L</th><th>Act</th></tr></thead><tbody>';
                mergedPositions.forEach(t => {
                    const st = (t.status || '').toLowerCase();
                    let badge = 'badge-open';
                    let stLabel = t.status || 'ACTIVE';
                    if (st === 'sl_hit') { badge = 'badge-loss'; stLabel = 'SL HIT'; }
                    else if (st === 'target_hit') { badge = 'badge-profit'; stLabel = 'TARGET'; }
                    else if (st === 'exited') { badge = 'badge-closed'; stLabel = 'EXITED'; }
                    const tokenKey = t.token || t.contract || t.symbol;
                    const fallbackLtp = tokenKey ? (ltpData[tokenKey] || ltpData[t.contract] || ltpData[t.symbol] || '') : '';
                    const displayLtp = (t.ltp !== undefined && t.ltp !== null && t.ltp > 0) ? t.ltp : fallbackLtp;
                    let pnlStr = '';
                    if (t.pnl !== undefined && t.pnl !== null && t.quantity > 0) {
                        const rawPnl = parseFloat(t.pnl);
                        const pctPnl = (t.entry_spot && t.entry_spot > 0) ? ((displayLtp - t.entry_spot) / t.entry_spot * 100).toFixed(2) : '0.00';
                        pnlStr = `${rawPnl >= 0 ? '+' : ''}${rawPnl.toFixed(2)} (${pctPnl}%)`;
                    } else if (displayLtp && t.entry_spot && t.entry_spot > 0) {
                        const rawPnl = (displayLtp - t.entry_spot) * (t.quantity || 1);
                        const pctPnl = ((displayLtp - t.entry_spot) / t.entry_spot * 100).toFixed(2);
                        pnlStr = `${rawPnl >= 0 ? '+' : ''}${rawPnl.toFixed(2)} (${pctPnl}%)`;
                    } else if (t.pnl_percent !== undefined && t.pnl_percent !== null) {
                        pnlStr = `${t.pnl_percent}%`;
                    } else if (t.pnl !== undefined && t.pnl !== null) {
                        pnlStr = `${t.pnl >= 0 ? '+' : ''}${parseFloat(t.pnl).toFixed(2)}`;
                    }
                    const pnlBadge = pnlStr ? (pnlStr.includes('-') ? 'badge-loss' : 'badge-profit') : '';
                    const qty = t.quantity || '';
                    const slVal = t.current_sl !== undefined && t.current_sl !== null ? t.current_sl : '';
                    const t1v = t.t1 !== undefined && t.t1 !== null ? t.t1 : '';
                    const t2v = t.t2 !== undefined && t.t2 !== null ? t.t2 : '';
                    const t3v = t.t3 !== undefined && t.t3 !== null ? t.t3 : '';
                    const entryVal = t.entry_spot !== undefined && t.entry_spot !== null ? t.entry_spot : '';
                    const uid = 'pos_' + (t.symbol || 'no_sym') + '_' + (t.source || '');
                    const es = editStates[uid];
                    let slCell, t1Cell, t2Cell, t3Cell, actCell;
                    if (es && es.active) {
                        const eng2 = t.engine === 'Index' ? 'index' : 'nifty50';
                        slCell = `<td><input id="sl_${uid}" value="${es.sl}" style="width:60px" oninput="editStates['${uid}'].sl=this.value" onchange="saveEdit('${uid}','${t.contract||t.symbol||''}','${eng2}')"></td>`;
                        t1Cell = `<td><input id="t1_${uid}" value="${es.t1}" style="width:60px" oninput="editStates['${uid}'].t1=this.value" onchange="saveEdit('${uid}','${t.contract||t.symbol||''}','${eng2}')"></td>`;
                        t2Cell = `<td><input id="t2_${uid}" value="${es.t2}" style="width:60px" oninput="editStates['${uid}'].t2=this.value" onchange="saveEdit('${uid}','${t.contract||t.symbol||''}','${eng2}')"></td>`;
                        t3Cell = `<td><input id="t3_${uid}" value="${es.t3}" style="width:60px" oninput="editStates['${uid}'].t3=this.value" onchange="saveEdit('${uid}','${t.contract||t.symbol||''}','${eng2}')"></td>`;
                        actCell = `<td><button class="btn-edit-save" onclick="saveEdit('${uid}','${t.contract||t.symbol||''}','${eng2}')">Save</button><button class="btn-edit-cancel" onclick="cancelEdit('${uid}')">X</button></td>`;
                    } else {
                        slCell = `<td>${slVal}</td>`;
                        t1Cell = `<td>${t1v}</td>`;
                        t2Cell = `<td>${t2v}</td>`;
                        t3Cell = `<td>${t3v}</td>`;
                        const canEdit = st === 'active';
                        if (canEdit) {
                            const eng2 = t.engine === 'Index' ? 'index' : 'nifty50';
                            actCell = `<td><button class="btn-edit" onclick="editRow('${uid}','${t.contract||t.symbol||''}','${slVal}','${t1v}','${t2v}','${t3v}')">Edit</button><button class="btn-exit" style="background:#da3633;color:#fff;border:none;padding:2px 8px;border-radius:4px;font-size:10px;cursor:pointer;margin-left:4px;font-weight:600;" onclick="manualExitPosition(this,'${t.contract||t.symbol||''}','${eng2}')">EXIT</button></td>`;
                        } else {
                            actCell = `<td></td>`;
                        }
                    }
                    posHtml += `<tr><td><strong>${t.symbol}</strong></td><td>${t.source}</td><td><span class="badge badge-open">${t.pattern||''}</span></td><td>${entryVal}</td>${slCell}${t1Cell}${t2Cell}${t3Cell}<td>${displayLtp || '-'}</td><td>${qty}</td><td><span class="badge ${badge}">${stLabel}</span></td><td>${pnlStr !== '' ? `<span class="badge ${pnlBadge}">${pnlStr}</span>` : '-'}</td>${actCell}</tr>`;
                });
                posHtml += '</tbody></table>';
            } else {
                posHtml = '<div class="empty-state">No positions found</div>';
            }
            document.getElementById('active-positions-body').innerHTML = posHtml;

            let jHtml = '';
            let filteredJournal = [...(journal || [])];
            const todayStr = new Date().toISOString().split('T')[0];
            allTrades.forEach(t => {
                const st = (t.status || '').toLowerCase();
                if (st === 'sl_hit' || st === 'target_hit' || st === 'exited') return;
                const et = (t.created_at || t.entry_time || '');
                const entryDate = et.split(' ')[0].split('T')[0];
                if (entryDate && entryDate < todayStr) {
                    filteredJournal.push({
                        Timestamp: et,
                        Symbol: t.contract || t.symbol,
                        Pattern: t.pattern || 'CARRY_FORWARD',
                        Action: 'CARRY_FORWARD',
                        Status: 'ACTIVE',
                        Entry: t.entry_spot,
                        SL: t.current_sl,
                        Target: t.t1,
                        RR: t.rr || '-',
                        'P&L %': t.pnl_percent || '-'
                    });
                }
            });

            if (journalFilter !== 'all') {
                filteredJournal = filteredJournal.filter(j => {
                    const sym = (j.Symbol || '').toUpperCase();
                    const isIndex = sym.includes('NIFTY') || sym.includes('BANK');
                    return journalFilter === 'index' ? isIndex : !isIndex;
                });
            }
            if (filteredJournal.length) {
                jHtml = '<table><thead><tr><th onclick="sortTable(this,0)">Time</th><th onclick="sortTable(this,1)">Symbol</th><th onclick="sortTable(this,2)">Pattern</th><th onclick="sortTable(this,3)">Action</th><th onclick="sortTable(this,4)">Status</th><th onclick="sortTable(this,5)">Entry</th><th onclick="sortTable(this,6)">SL</th><th onclick="sortTable(this,7)">Target</th><th onclick="sortTable(this,8)">RR</th><th onclick="sortTable(this,9)">P&L %</th></tr></thead><tbody>';
                filteredJournal.forEach(j => {
                    const pnlv = j['P&L %'] || '-';
                    const ts = j.Timestamp || '';
                    let time = '';
                    if (ts) {
                        const p = ts.split(' ');
                        const dp = p[0] ? p[0].split('-') : [];
                        const tp = p[1] ? p[1].split(':') : [];
                        if (dp.length === 3 && tp.length === 3)
                            time = `${dp[2]}:${dp[1]}:${dp[0].slice(-2)}-${tp[0]}:${tp[1]}`;
                    }
                    const act = j.Action || '';
                    const st = j.Status || '';
                    let badge = 'badge-open';
                    if (act.startsWith('EXIT')) badge = 'badge-closed';
                    else if (st === 'FAILED') badge = 'badge-failed';
                    else if (st === 'MUTATED') badge = 'badge-mutated';
                    let pnlBadge = '';
                    if (pnlv !== '-') {
                        const pn = parseFloat(pnlv);
                        pnlBadge = pn >= 0 ? 'badge-profit' : 'badge-loss';
                    }
                    const entryVal = j['Entry'] || '';
                    const slVal = j['SL'] || '';
                    const targetVal = j['Target'] || '';
                    const rrVal = j['RR'] || '';
                    jHtml += `<tr><td style="font-size:11px">${time}</td><td>${j.Symbol||''}</td><td>${j.Pattern||''}</td><td><span class="badge ${badge}">${act}</span></td><td>${st}</td><td>${entryVal}</td><td>${slVal}</td><td>${targetVal}</td><td>${rrVal}</td><td><span class="badge ${pnlBadge}">${pnlv}</span></td></tr>`;
                });
                jHtml += '</tbody></table>';
            } else {
                jHtml = '<p class="empty-state">No journal entries yet</p>';
            }
            document.getElementById('journal-body').innerHTML = jHtml;
            renderScanTab();

            let logHtml = '';
            let allLogs = [];
            if (logFilter === 'all' || logFilter === 'index') {
                allLogs = allLogs.concat(d.programs?.index?.log_tail || []);
            }
            if (logFilter === 'all' || logFilter === 'nifty50') {
                allLogs = allLogs.concat(d.programs?.nifty50?.log_tail || []);
            }
            allLogs.forEach(l => { logHtml += '<div>'+l.replace(/</g,'&lt;').replace(/>/g,'&gt;')+'</div>'; });
            document.getElementById('log-body').innerHTML = logHtml;
            const logBox = document.getElementById('log-body');
            logBox.scrollTop = logBox.scrollHeight;

            refreshLiveExec();
            document.getElementById('last-updated').textContent = 'Last refreshed: ' + new Date().toLocaleString();
            reapplySorts();
        }

        function renderBacktest() {
            const d = window._lastData;
            if (!d) return;
            const journal = d.journal || [];
            if (!journal.length) {
                document.getElementById('backtest-body').innerHTML = '<p class="empty-state">No journal data to analyze</p>';
                return;
            }
            const stats = {};
            journal.forEach(j => {
                const sym = j.Symbol || 'UNKNOWN';
                const act = j.Action || '';
                const pnlStr = j['P&L %'] || '-';
                const pnl = parseFloat(pnlStr.replace('%', '')) || 0;
                const entry = j.Entry || '';
                const sl = j.SL || '';
                const target = j.Target || '';
                const rr = j.RR || '';
                if (!stats[sym]) stats[sym] = { entries: 0, slHits: 0, targetHits: 0, totalPnl: 0, trades: 0, rrSum: 0, rrCount: 0 };
                if (act.startsWith('BACKTEST') || act.startsWith('BUY') || act.startsWith('ENTRY')) {
                    stats[sym].entries++;
                    stats[sym].trades++;
                    if (rr && !isNaN(parseFloat(rr))) { stats[sym].rrSum += parseFloat(rr); stats[sym].rrCount++; }
                }
                if (act === 'EXIT_SL') stats[sym].slHits++;
                if (act === 'EXIT_T3') stats[sym].targetHits++;
                if (pnl && pnlStr !== '-') stats[sym].totalPnl += pnl;
            });
            let html = '<table><thead><tr><th onclick="sortTable(this,0)">Symbol</th><th onclick="sortTable(this,1)">Entries</th><th onclick="sortTable(this,2)">SL Hits</th><th onclick="sortTable(this,3)">Target Hits</th><th onclick="sortTable(this,4)">Win Rate</th><th onclick="sortTable(this,5)">Total P&L</th><th onclick="sortTable(this,6)">Avg P&L</th><th onclick="sortTable(this,7)">Avg RR</th></tr></thead><tbody>';
            Object.entries(stats).sort((a, b) => b[1].entries - a[1].entries).forEach(([sym, s]) => {
                const exits = s.slHits + s.targetHits;
                const wr = exits > 0 ? ((s.targetHits / exits) * 100).toFixed(1) : '-';
                const avgPnl = s.trades > 0 ? (s.totalPnl / s.trades).toFixed(2) : '-';
                const avgRR = s.rrCount > 0 ? (s.rrSum / s.rrCount).toFixed(2) : '-';
                html += `<tr><td><strong>${sym}</strong></td><td>${s.entries}</td><td>${s.slHits}</td><td>${s.targetHits}</td><td>${wr}${wr !== '-' ? '%' : ''}</td><td class="${s.totalPnl >= 0 ? 'pnl-positive' : 'pnl-negative'}">${s.totalPnl.toFixed(2)}%</td><td>${avgPnl}%</td><td>${avgRR}</td></tr>`;
            });
            html += '</tbody></table>';
            document.getElementById('backtest-body').innerHTML = html;
            reapplySorts();
        }

        async function refreshData() {
            try {
                const r = await fetch('/api/status');
                const d = await r.json();
                window._lastData = d;

                for (const [pid, prog] of Object.entries(d.programs || {})) {
                    const btn = document.querySelector(`.prog-card[data-prog="${pid}"]`);
                    if (!btn) continue;
                    const dot = btn.querySelector('.status-dot');
                    const label = btn.querySelector('.status-label');
                    const bar = btn.querySelector('.status-bar');
                    const startBtn = btn.querySelector('.start-btn');
                    const stopBtn = btn.querySelector('.stop-btn');
                    if (prog.running) {
                        dot.className = 'status-dot live';
                        label.textContent = 'Live';
                        label.className = 'status-label live';
                        bar.className = 'status-bar live';
                        startBtn.disabled = true;
                        startBtn.style.opacity = '0.4';
                        stopBtn.disabled = false;
                        stopBtn.style.opacity = '1';
                    } else {
                        dot.className = 'status-dot closed';
                        label.textContent = 'Closed';
                        label.className = 'status-label closed';
                        bar.className = 'status-bar closed';
                        startBtn.disabled = false;
                        startBtn.style.opacity = '1';
                        stopBtn.disabled = true;
                        stopBtn.style.opacity = '0.4';
                    }
                }

                const cfg = d.config || {};
                for (const [pid, progCfg] of Object.entries(cfg)) {
                    const inputs = document.querySelectorAll(`.config-input[data-prog="${pid}"]`);
                    inputs.forEach(inp => {
                        const field = inp.getAttribute('data-field');
                        if (progCfg[field] !== undefined) inp.value = progCfg[field];
                    });
                }

                for (const [pid, prog] of Object.entries(d.programs || {})) {
                    (prog.log_tail || []).forEach(line => {
                        if (!line.includes('[ERROR]')) return;
                        const key = pid + line.slice(0, 60);
                        if (seenAlerts.has(key)) return;
                        seenAlerts.add(key);
                        if (seenAlerts.size > 200) seenAlerts.clear();
                        showAlert(pid + ': ' + line.split('[ERROR]')[1] || line);
                    });
                }

                let logHtml = '';
                let allLogs = [];
                const lf = logFilter || 'all';
                if (lf === 'all' || lf === 'index') {
                    allLogs = allLogs.concat(d.programs?.index?.log_tail || []);
                }
                if (lf === 'all' || lf === 'nifty50') {
                    allLogs = allLogs.concat(d.programs?.nifty50?.log_tail || []);
                }
                allLogs.forEach(l => { logHtml += '<div>'+l.replace(/</g,'&lt;').replace(/>/g,'&gt;')+'</div>'; });
                document.getElementById('log-body').innerHTML = logHtml || '<p class="empty-state">Waiting for log data...</p>';
                const logBox = document.getElementById('log-body');
                if (logBox) logBox.scrollTop = logBox.scrollHeight;

                for (const pid of ['index', 'nifty50']) {
                    const enabled = pid === 'index' ? d.live_execution_index : d.live_execution;
                    const sw = document.getElementById('live-toggle-' + pid);
                    const lb = document.getElementById('live-label-' + pid);
                    if (sw) { sw.classList.toggle('on', !!enabled); }
                    if (lb) { lb.textContent = enabled ? 'LIVE' : 'SCAN'; lb.style.color = enabled ? '#3fb950' : '#8b949e'; }
                }
                renderScanTab();
                renderReport();
                refreshTokenStatus();
                refreshBacktestMode();
            } catch(e) { console.log('Refresh error:', e); }
        }

        async function refreshTokenStatus() {
            // TradingView Open-Source Edition - Free data feed active
            return;
        }

        async function refreshBacktestMode() {
            try {
                const r = await fetch('/api/backtest/mode');
                const d = await r.json();
                const btToggle = document.getElementById('backtest-toggle');
                if (btToggle) btToggle.checked = d.enabled === true;
                if (btToggle && btToggle.checked) renderBacktest();
            } catch(e) { console.log('Backtest mode error:', e); }
        }

        async function showTokenPanel() {
            document.getElementById('token-panel').style.display = 'block';
            document.getElementById('token-gen-btn').style.display = 'none';
            document.getElementById('token-feedback').textContent = '';
            try {
                const r = await fetch('/api/token/url');
                const d = await r.json();
                document.getElementById('token-url-text').textContent = d.url || 'Error loading URL';
            } catch(e) {
                document.getElementById('token-url-text').textContent = 'Failed to load login URL';
            }
        }

        function hideTokenPanel() {
            document.getElementById('token-panel').style.display = 'none';
            document.getElementById('token-gen-btn').style.display = 'inline-block';
        }

        function copyTokenUrl() {
            const url = document.getElementById('token-url-text').textContent;
            if (url && url.startsWith('http')) {
                navigator.clipboard.writeText(url).then(() => {
                    const hint = document.querySelector('.token-copy-hint');
                    hint.textContent = 'Copied!';
                    setTimeout(() => { hint.textContent = 'Click to copy'; }, 2000);
                }).catch(() => {});
            }
        }

        async function submitToken() {
            const input = document.getElementById('token-redirect-input');
            const fb = document.getElementById('token-feedback');
            const raw = input.value.trim();
            if (!raw) { fb.textContent = 'Please paste the redirect URL'; fb.style.color = '#f85149'; return; }
            let requestToken = raw;
            if (requestToken.includes('request_token=')) {
                requestToken = requestToken.split('request_token=')[1].split('&')[0];
            }
            fb.textContent = 'Exchanging token...';
            fb.style.color = '#8b949e';
            try {
                const r = await fetch('/api/token/exchange', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({request_token: requestToken})
                });
                const d = await r.json();
                if (d.ok) {
                    fb.textContent = 'Token saved successfully!';
                    fb.style.color = '#3fb950';
                    document.getElementById('token-gen-btn').style.display = 'none';
                    setTimeout(() => { hideTokenPanel(); refreshData(); }, 1500);
                } else {
                    fb.textContent = 'Failed: ' + (d.error || 'Unknown error');
                    fb.style.color = '#f85149';
                }
            } catch(e) {
                fb.textContent = 'Error: ' + e.message;
                fb.style.color = '#f85149';
            }
        }

        let seenAlerts = new Set();

        function showAlert(msg) {
            let el = document.getElementById('alert-toast');
            if (!el) {
                el = document.createElement('div');
                el.id = 'alert-toast';
                document.body.appendChild(el);
            }
            el.innerHTML = '<span class="alert-icon">&#9888;</span><span class="alert-msg">' + msg.replace(/</g,'&lt;').replace(/>/g,'&gt;') + '</span><button class="alert-close" onclick="this.parentElement.remove()">&times;</button>';
            el.className = 'alert-toast show';
            clearTimeout(el._hideTimer);
            el._hideTimer = setTimeout(() => { if (el) el.className = 'alert-toast'; }, 15000);
        }

        async function monthlyExport() {
            try {
                const r = await fetch('/api/export/monthly', {method: 'POST'});
                const d = await r.json();
                const fb = document.getElementById('cfg-fb-backtest');
                if (d.ok) {
                    fb.textContent = d.exported > 0 ? `Exported ${d.exported} trades to ${d.sheets.join(', ')}` : 'No completed trades to export';
                    fb.style.color = '#3fb950';
                } else {
                    fb.textContent = 'Export failed';
                    fb.style.color = '#f85149';
                }
                setTimeout(() => { fb.textContent = ''; }, 5000);
                if (d.ok) setTimeout(refreshData, 500);
            } catch(e) { console.log(e); }
        }

        async function clearLogs() {
            try {
                const r = await fetch('/api/logs/clear', {method: 'POST'});
                const d = await r.json();
                if (d.ok) setTimeout(refreshData, 300);
            } catch(e) { console.log(e); }
        }

        async function clearJournal() {
            try {
                const r = await fetch('/api/journal/clear', {method: 'POST'});
                const d = await r.json();
                if (d.ok) setTimeout(refreshData, 300);
            } catch(e) { console.log(e); }
        }

        setInterval(refreshData, {{ refresh * 1000 }});
        window.addEventListener('load', () => { refreshData(); });
    </script>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Segoe UI', sans-serif; background: #0d1117; color: #c9d1d9; padding: 20px; }
        h1 { color: #58a6ff; margin-bottom: 16px; font-size: 22px; display: flex; align-items: center; gap: 10px; }
        h1 small { font-size: 12px; color: #8b949e; font-weight: normal; }
        h2 { color: #8b949e; font-size: 14px; margin-bottom: 10px; border-bottom: 1px solid #30363d; padding-bottom: 6px; text-transform: uppercase; letter-spacing: 0.5px; }

        .empty-state { color: #8b949e; padding: 20px; text-align: center; font-size: 13px; }

        .token-banner { border-radius: 8px; padding: 10px 14px; margin-bottom: 16px; font-size: 13px; transition: all 0.3s; }
        .token-banner.token-hidden { display: none; }
        .token-banner.token-valid { background: #3fb95022; border: 1px solid #3fb950; color: #3fb950; }
        .token-banner.token-expired { background: #d2992222; border: 1px solid #d29922; color: #d29922; }
        .token-banner.token-missing { background: #f8514922; border: 1px solid #f85149; color: #f85149; }
        .token-banner-row { display: flex; align-items: center; justify-content: space-between; gap: 12px; }
        .token-gen-btn { background: #238636; color: #fff; border: none; border-radius: 4px; padding: 5px 14px; font-size: 12px; cursor: pointer; font-weight: 600; }
        .token-gen-btn:hover { background: #2ea043; }
        .token-gen-btn:disabled { opacity: 0.5; cursor: not-allowed; }

        .token-panel { margin-top: 12px; padding: 14px; background: #0d1117; border-radius: 6px; border: 1px solid #30363d; }
        .token-step { display: flex; align-items: center; gap: 8px; font-size: 12px; color: #c9d1d9; margin-bottom: 8px; }
        .token-step-num { background: #21262d; color: #58a6ff; width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: bold; flex-shrink: 0; }
        .token-url-box { background: #161b22; border: 1px solid #30363d; border-radius: 4px; padding: 8px 10px; font-size: 11px; color: #58a6ff; word-break: break-all; cursor: pointer; margin: 6px 0 12px; display: flex; justify-content: space-between; align-items: center; gap: 8px; }
        .token-url-box:hover { border-color: #58a6ff; }
        .token-copy-hint { font-size: 10px; color: #8b949e; white-space: nowrap; }
        .token-input-row { display: flex; gap: 8px; margin: 8px 0; }
        .token-input { flex: 1; background: #161b22; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 6px 10px; font-size: 12px; }
        .token-input:focus { outline: none; border-color: #58a6ff; }
        .token-submit-btn { background: #238636; color: #fff; border: none; border-radius: 4px; padding: 6px 14px; font-size: 12px; cursor: pointer; font-weight: 600; white-space: nowrap; }
        .token-submit-btn:hover { background: #2ea043; }
        .token-feedback { font-size: 12px; margin: 6px 0; }
        .token-close-btn { background: transparent; color: #8b949e; border: 1px solid #30363d; border-radius: 4px; padding: 3px 10px; font-size: 11px; cursor: pointer; margin-top: 6px; }
        .token-close-btn:hover { color: #c9d1d9; }

        .stats-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-bottom: 20px; }
        .stat-card { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 14px; text-align: center; }
        .stat-card .value { font-size: 26px; font-weight: bold; }
        .stat-card .label { font-size: 11px; color: #8b949e; margin-top: 2px; text-transform: uppercase; }

        .program-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 20px; }
        .prog-card { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 14px; position: relative; overflow: hidden; }

        .status-bar { position: absolute; top: 0; left: 0; width: 4px; height: 100%; transition: background 0.3s; }
        .status-bar.live { background: #3fb950; }
        .status-bar.closed { background: #f85149; }

        .prog-header { display: flex; align-items: center; gap: 10px; margin-bottom: 6px; }
        .prog-icon { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }
        .prog-name { font-weight: 600; font-size: 13px; flex: 1; }
        .prog-live-toggle { display: flex; align-items: center; gap: 4px; cursor: pointer; padding: 2px 6px; border-radius: 4px; user-select: none; }
        .prog-live-toggle:hover { background: #1c2128; }
        .live-label { font-size: 10px; font-weight: 600; color: #8b949e; min-width: 28px; text-align: center; }
        .live-switch { width: 24px; height: 12px; background: #30363d; border-radius: 6px; position: relative; transition: background 0.2s; }
        .live-switch::after { content: ''; position: absolute; top: 1px; left: 1px; width: 10px; height: 10px; background: #8b949e; border-radius: 50%; transition: all 0.2s; }
        .live-switch.on { background: #238636; }
        .live-switch.on::after { left: 13px; background: #fff; }
        .live-switch.on + .live-label { color: #3fb950; }
        .prog-desc { font-size: 11px; color: #8b949e; margin-bottom: 10px; }
        .prog-footer { display: flex; align-items: center; justify-content: space-between; }

        .status-group { display: flex; align-items: center; gap: 5px; }
        .status-dot { display: inline-block; width: 8px; height: 8px; border-radius: 50%; }
        .status-dot.live { background: #3fb950; box-shadow: 0 0 8px #3fb950aa; animation: pulse 1.5s infinite; }
        .status-dot.closed { background: #f85149; }
        @keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.5; } }
        .status-label { font-size: 11px; font-weight: 600; }
        .status-label.live { color: #3fb950; }
        .status-label.closed { color: #f85149; }

        .prog-actions { display: flex; gap: 6px; }
        .prog-actions button { padding: 4px 14px; border-radius: 6px; cursor: pointer; font-size: 11px; font-weight: 600; border: 1px solid; transition: all 0.2s; }
        .prog-actions button:disabled { cursor: not-allowed; }
        .start-btn { background: #3fb95022; border-color: #3fb950; color: #3fb950; }
        .start-btn:hover:not(:disabled) { background: #3fb95044; }
        .stop-btn { background: #f8514922; border-color: #f85149; color: #f85149; }
        .stop-btn:hover:not(:disabled) { background: #f8514944; }
        .anchor-btn { background: #d2992222; border-color: #d29922; color: #d29922; }
        .anchor-btn:hover:not(:disabled) { background: #d2992244; }
        .anchor-btn:disabled { opacity: 0.5; cursor: not-allowed; }

        .prog-config { border-top: 1px solid #30363d; margin-top: 0; }
        .config-header { padding: 6px 14px; font-size: 11px; color: #8b949e; cursor: pointer; user-select: none; display: flex; align-items: center; gap: 6px; }
        .config-header:hover { color: #c9d1d9; background: #1c2128; }
        .config-arrow { font-size: 9px; }
        .config-body { display: none; padding: 6px 14px 10px; background: #0d1117; }
        .config-row { display: flex; align-items: center; justify-content: space-between; padding: 3px 0; gap: 8px; }
        .config-label { font-size: 10px; color: #8b949e; }
        .config-input { background: #161b22; color: #c9d1d9; border: 1px solid #30363d; border-radius: 3px; padding: 2px 6px; font-size: 10px; width: 120px; }
        .config-input:focus { outline: none; border-color: #58a6ff; }
        .config-feedback { font-size: 10px; margin-left: 8px; }

        .section-panel { background: #161b22; border: 1px solid #30363d; border-radius: 8px; margin-bottom: 14px; overflow: hidden; }
        .section-panel .section-header { background: #1c2128; color: #c9d1d9; font-size: 12px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; padding: 8px 14px; border-bottom: 1px solid #30363d; display: flex; align-items: center; justify-content: space-between; gap: 10px; }
        .section-panel .empty-state { color: #8b949e; padding: 20px; text-align: center; font-size: 13px; }
        .filter-select { background: #0d1117; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 3px 8px; font-size: 11px; cursor: pointer; }
        .filter-select:focus { outline: none; border-color: #58a6ff; }
        .clear-logs-btn { background: #21262d; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 3px 10px; font-size: 10px; cursor: pointer; white-space: nowrap; }
        .clear-logs-btn:hover { background: #30363d; }
        .pos-filter-btn { background: #21262d; color: #8b949e; border: 1px solid #30363d; border-radius: 4px; padding: 2px 8px; font-size: 10px; cursor: pointer; white-space: nowrap; }
        .pos-filter-btn:hover { color: #c9d1d9; }
        .pos-filter-btn.active { background: #1f6feb33; color: #58a6ff; border-color: #58a6ff; }
        .export-btn { background: #21262d; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 3px 10px; font-size: 10px; cursor: pointer; white-space: nowrap; }
        .export-btn:hover { background: #30363d; }

        .toggle-label { display: inline-flex; align-items: center; gap: 0; cursor: pointer; user-select: none; font-size: 10px; color: #8b949e; }
        .toggle-label input { display: none; }
        .toggle-slider { width: 28px; height: 14px; background: #30363d; border-radius: 10px; position: relative; transition: background 0.2s; }
        .toggle-slider::after { content: ''; position: absolute; top: 2px; left: 2px; width: 10px; height: 10px; background: #8b949e; border-radius: 50%; transition: all 0.2s; }
        .toggle-label input:checked + .toggle-slider { background: #238636; }
        .toggle-label input:checked + .toggle-slider::after { left: 16px; background: #fff; }

        .reports-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; align-items: start; }
        .reports-left .section-panel { margin-bottom: 0; border-top: none; border-radius: 0 0 8px 8px; }

        .pnl-positive { color: #3fb950 !important; font-weight: 600; }
        .pnl-negative { color: #f85149 !important; font-weight: 600; }
        .reports-right .section-panel { margin-bottom: 0; border-top: none; border-radius: 0 0 8px 8px; }
        .left-tab-bar { display: flex; gap: 0; background: #1c2128; border: 1px solid #30363d; border-bottom: none; border-radius: 8px 8px 0 0; overflow: hidden; }
        .left-tab-btn { background: transparent; border: none; color: #8b949e; padding: 8px 16px; cursor: pointer; font-size: 12px; font-weight: 600; border-bottom: 2px solid transparent; flex: 1; text-align: center; }
        .left-tab-btn:hover { color: #c9d1d9; background: #21262d; }
        .left-tab-btn.active { color: #58a6ff; border-bottom-color: #58a6ff; background: #161b22; }
        .left-tab-content { display: none; }
        .left-tab-content.active { display: block; }

        .tab-bar { display: flex; gap: 0; background: #1c2128; border: 1px solid #30363d; border-bottom: none; border-radius: 8px 8px 0 0; overflow: hidden; }
        .tab-btn { background: transparent; border: none; color: #8b949e; padding: 8px 16px; cursor: pointer; font-size: 12px; font-weight: 600; border-bottom: 2px solid transparent; flex: 1; text-align: center; }
        .tab-btn:hover { color: #c9d1d9; background: #21262d; }
        .tab-btn.active { color: #58a6ff; border-bottom-color: #58a6ff; background: #161b22; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }

        table { width: 100%; border-collapse: collapse; margin-bottom: 16px; background: #161b22; border-radius: 8px; overflow: hidden; }
        th { background: #21262d; color: #8b949e; font-size: 11px; text-transform: uppercase; padding: 8px 10px; text-align: left; border-bottom: 1px solid #30363d; cursor: pointer; user-select: none; }
        th:hover { color: #58a6ff; }
        th::after { content: ' \\25B4\\25BE'; font-size: 8px; opacity: 0.3; }
        td { padding: 7px 10px; border-bottom: 1px solid #21262d; font-size: 12px; }
        tr:hover td { background: #1c2128; }

        .badge { display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 10px; font-weight: 600; }
        .badge-open { background: #1f6feb33; color: #58a6ff; }
        .badge-closed { background: #3fb95033; color: #3fb950; }
        .badge-failed { background: #f8514933; color: #f85149; }
        .badge-mutated { background: #d2992233; color: #d29922; }
        .badge-profit { background: #3fb95033; color: #3fb950; }
        .badge-loss { background: #f8514933; color: #f85149; }

        .log-box { background: #0d1117; border: 1px solid #30363d; border-radius: 8px; padding: 10px; font-family: 'Consolas', monospace; font-size: 11px; max-height: 400px; overflow-y: auto; line-height: 1.5; }
        .log-box div:nth-child(odd) { background: #161b22; }
        .match-highlight { background: #1f6feb11; border-left: 3px solid #58a6ff; padding: 2px 8px; margin: 2px 0; font-family: 'Consolas', monospace; font-size: 11px; }
        .toggle-wrap { display: flex; align-items: center; gap: 8px; font-size: 12px; }
        .toggle-switch { position: relative; width: 40px; height: 20px; background: #30363d; border-radius: 10px; cursor: pointer; transition: 0.2s; }
        .toggle-switch.active { background: #3fb950; }
        .toggle-switch::after { content: ''; position: absolute; width: 16px; height: 16px; border-radius: 50%; background: #fff; top: 2px; left: 2px; transition: 0.2s; }
        .toggle-switch.active::after { left: 22px; }
        .scan-section-title { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; color: #8b949e; padding: 8px 14px 4px; }
        .btn-edit { background: #1f6feb33; color: #58a6ff; border: 1px solid #1f6feb66; border-radius: 4px; padding: 2px 8px; font-size: 10px; cursor: pointer; }
        .btn-edit:hover { background: #1f6feb66; }
        .btn-edit-save { background: #3fb95033; color: #3fb950; border: 1px solid #3fb95066; border-radius: 4px; padding: 2px 8px; font-size: 10px; cursor: pointer; margin-right: 4px; }
        .btn-edit-save:hover { background: #3fb95066; }
        .btn-edit-cancel { background: #f8514933; color: #f85149; border: 1px solid #f8514966; border-radius: 4px; padding: 2px 8px; font-size: 10px; cursor: pointer; }
        .btn-edit-cancel:hover { background: #f8514966; }
        .toast-container { position: fixed; top: 16px; right: 16px; z-index: 9999; display: flex; flex-direction: column; gap: 8px; }
        .toast { padding: 10px 16px; border-radius: 8px; font-size: 13px; color: #fff; box-shadow: 0 4px 12px rgba(0,0,0,0.3); animation: toast-in 0.3s ease; }
        .toast-success { background: #3fb950; }
        .toast-error { background: #f85149; }
        @keyframes toast-in { from { opacity: 0; transform: translateX(40px); } to { opacity: 1; transform: translateX(0); } }

        .alert-toast { position: fixed; top: 12px; right: 12px; z-index: 9999; max-width: 420px; background: #f85149; color: #fff; padding: 10px 14px; border-radius: 8px; font-size: 12px; display: flex; align-items: flex-start; gap: 8px; box-shadow: 0 4px 20px rgba(0,0,0,0.5); transform: translateX(120%); opacity: 0; transition: all 0.4s; }
        .alert-toast.show { transform: translateX(0); opacity: 1; }
        .alert-icon { font-size: 16px; flex-shrink: 0; margin-top: 1px; }
        .alert-msg { flex: 1; word-break: break-word; }
        .alert-close { background: none; border: none; color: rgba(255,255,255,0.7); font-size: 18px; cursor: pointer; padding: 0 0 0 4px; line-height: 1; flex-shrink: 0; }
        .alert-close:hover { color: #fff; }
        .last-updated { color: #8b949e; font-size: 11px; text-align: right; margin-top: 8px; }

        @media (max-width: 1100px) {
            .reports-grid { grid-template-columns: 1fr; }
        }
        @media (max-width: 900px) {
            .stats-grid { grid-template-columns: repeat(2, 1fr); }
            .program-grid { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
    <div class="toast-container" id="toast-container"></div>
    <h1>Trading Control Center <small>Steering & Dashboard</small></h1>

    <div style="background:#161b22;border:1px solid #30363d;border-radius:8px;padding:10px 16px;margin-bottom:16px;display:flex;align-items:center;justify-space-between;">
        <div style="display:flex;align-items:center;gap:10px;">
            <span style="background:#238636;color:#ffffff;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:bold;">● Free Data Feed Active</span>
            <span style="font-size:12px;color:#c9d1d9;">TradingView & Yahoo Finance Open-Source Edition <span style="color:#8b949e;">(Zero API Key / Credentials Required)</span></span>
        </div>
        <span style="font-size:11px;color:#8b949e;font-weight:600;">Port: 6060</span>
    </div>

    <div class="stats-grid">
        <div class="stat-card">
            <div class="value" id="stat-active" style="color:#58a6ff;">0</div>
            <div class="label">Active Positions</div>
        </div>
        <div class="stat-card">
            <div class="value" id="stat-total" style="color:#58a6ff;">0</div>
            <div class="label">Total Trades</div>
        </div>
        <div class="stat-card">
            <div class="value" id="stat-winrate" style="color:#8b949e;">0%</div>
            <div class="label">Win Rate</div>
        </div>
        <div class="stat-card">
            <div class="value" id="stat-pnl" style="color:#8b949e;">0%</div>
            <div class="label">P&L</div>
        </div>
    </div>

    <h2>Programs</h2>
    <div class="program-grid">
        {% for pid, prog in programs.items() %}
        <div class="prog-card" data-prog="{{ pid }}">
            <div class="status-bar closed"></div>
            <div class="prog-header">
                <span class="prog-icon" style="background:{{ prog.color }}"></span>
                <span class="prog-name">{{ prog.name }}</span>
                {% if pid != 'daily' %}
                <div class="prog-live-toggle" onclick="event.stopPropagation();toggleCardLive('{{ pid }}')">
                    <span id="live-label-{{ pid }}" class="live-label">SCAN</span>
                    <span id="live-toggle-{{ pid }}" class="live-switch"></span>
                </div>
                {% endif %}
                <div class="status-group">
                    <span class="status-dot closed"></span>
                    <span class="status-label closed">Closed</span>
                </div>
            </div>
            <div class="prog-desc">{{ prog.desc }}</div>
            <div class="prog-footer" style="margin-bottom:{% if prog.config_fields %}6px{% else %}0{% endif %};">
                <div class="prog-actions">
                    <button class="start-btn" onclick="event.stopPropagation();toggleProgram('{{ pid }}','start')">Start</button>
                    <button class="stop-btn" onclick="event.stopPropagation();toggleProgram('{{ pid }}','stop')" disabled>Stop</button>
                </div>
            </div>
            {% if prog.config_fields %}
            <div class="prog-config">
                <div class="config-header" onclick="event.stopPropagation();toggleConfig(this)">
                    <span class="config-arrow">&#9654;</span> Configuration
                </div>
                <div class="config-body">
                    {% for field_key, field in prog.config_fields.items() %}
                    <div class="config-row">
                        <span class="config-label">{{ field.label }}</span>
                        {% if field.type == "select" %}
                        <select class="config-input" data-prog="{{ pid }}" data-field="{{ field_key }}" onchange="event.stopPropagation();saveConfig('{{ pid }}')">
                            {% for opt in field.options %}
                            <option value="{{ opt }}">{{ opt }}</option>
                            {% endfor %}
                        </select>
                        {% else %}
                        <input type="number" class="config-input" data-prog="{{ pid }}" data-field="{{ field_key }}" value="{{ field.default }}" step="any" onchange="event.stopPropagation();saveConfig('{{ pid }}')">
                        {% endif %}
                    </div>
                    {% endfor %}
                    <span class="config-feedback" id="cfg-fb-{{ pid }}"></span>
                </div>
            </div>
            {% endif %}
        </div>
        {% endfor %}
    </div>

    <h2>Live Reports</h2>

    <div class="reports-grid">
        <div class="reports-left">
            <div class="left-tab-bar">
                <button class="left-tab-btn active" onclick="switchLeftTab('scan-tab-left')">Today's scan</button>
                <button class="left-tab-btn" onclick="switchLeftTab('tvchart-tab')">📈 TradingView Chart</button>
                <button class="left-tab-btn" onclick="switchLeftTab('backtest-tab')">Backtest Summary</button>
                <div style="display:flex;align-items:center;gap:6px;margin-left:auto;padding:0 10px;">
                    <label class="toggle-label" title="Toggle backtest mode">
                        <span style="font-size:10px;font-weight:normal;text-transform:none;letter-spacing:0;margin-right:6px;color:#8b949e;">Off</span>
                        <input type="checkbox" id="backtest-toggle" onchange="toggleBacktestMode(this.checked)">
                        <span class="toggle-slider"></span>
                        <span style="font-size:10px;font-weight:normal;text-transform:none;letter-spacing:0;margin-left:6px;color:#8b949e;">On</span>
                    </label>
                </div>
            </div>
            <div id="tvchart-tab" class="left-tab-content">
                <div class="section-panel" style="height:620px;">
                    <div class="section-header">
                        <span>TradingView Interactive Chart</span>
                        <div style="display:flex;align-items:center;gap:8px;">
                            <input type="text" id="tv-symbol-input" value="NIFTY" style="background:#161b22;border:1px solid #30363d;color:#c9d1d9;padding:4px 8px;border-radius:4px;font-size:11px;font-weight:bold;width:120px;" placeholder="Symbol">
                            <button onclick="loadTVChartFromInput()" style="background:#2962ff;color:#fff;border:none;padding:4px 10px;border-radius:4px;font-weight:bold;cursor:pointer;font-size:11px;">Load Chart 📈</button>
                        </div>
                    </div>
                    <div id="tv_chart_container" style="height:calc(100% - 40px);width:100%;">
                        <div style="padding:40px;text-align:center;color:#8b949e;">Click any symbol in Today's Scan table or enter a ticker above to load the interactive TradingView chart.</div>
                    </div>
                </div>
            </div>
            <div id="scan-tab-left" class="left-tab-content active">
                <div class="section-panel">
                    <div class="section-header">
                        <span>Positions</span>
                        <div style="display:flex;gap:4px;align-items:center;">
                            <button class="pos-filter-btn active" onclick="setFilter('position','active')">Active</button>
                            <button class="pos-filter-btn" onclick="setFilter('position','completed')">Completed</button>
                            <button class="pos-filter-btn" onclick="setFilter('position','sl_hit')">SL Hit</button>
                            <button class="pos-filter-btn" onclick="setFilter('position','all')">All</button>
                            <button class="btn-exit-all" onclick="manualExitAllPositions()" style="padding:2px 10px;background:#da3633;border:1px solid #f85149;color:#fff;border-radius:4px;font-size:10px;cursor:pointer;font-weight:600;margin-left:8px;">EXIT ALL</button>
                        </div>
                    </div>
                    <div id="active-positions-body"><p class="empty-state">No positions</p></div>
                </div>
            </div>
            <div id="scan-tab-left" class="left-tab-content">
                <div class="section-panel">
                    <div class="section-header">
                        <span>Trade Details</span>
                        <div style="display:flex;gap:12px;align-items:center">
                            <div class="toggle-wrap">
                                <span id="live-exec-label" style="color:#8b949e;font-size:11px">N50: OFF</span>
                                <div id="live-exec-toggle" class="toggle-switch" onclick="toggleLiveExecution('nifty50')"></div>
                            </div>
                            <div class="toggle-wrap">
                                <span id="live-exec-label-idx" style="color:#8b949e;font-size:11px">IDX: OFF</span>
                                <div id="live-exec-toggle-idx" class="toggle-switch" onclick="toggleLiveExecution('index')"></div>
                            </div>
                            <select id="scan-engine-filter" onchange="renderScanTab()" class="filter-select" style="width:auto">
                                <option value="all" selected>All Options</option>
                                <option value="index">Index Options</option>
                                <option value="nifty50">Stock Options</option>
                            </select>
                            <button class="btn-scan-clear" onclick="clearScanData()" style="padding:2px 10px;background:inherit;border:1px solid #f85149;color:#f85149;border-radius:4px;font-size:10px;cursor:pointer">Clear</button>
                            <button class="btn-scan-export" onclick="scanExport()" style="padding:2px 10px;background:inherit;border:1px solid #58a6ff;color:#58a6ff;border-radius:4px;font-size:10px;cursor:pointer">Export</button>
                        </div>
                    </div>
                    <div id="scan-body"><p class="empty-state">No scan trades yet</p></div>
                </div>
            </div>
            <div id="backtest-tab" class="left-tab-content">
                <div class="section-panel">
                    <div class="section-header">
                        <span>Backtest Summary</span>
                        <div style="display:flex;align-items:center;gap:8px;">
                            <button class="export-btn" onclick="monthlyExport()">Export Month</button>
                        </div>
                    </div>
                    <div id="backtest-body"><p class="empty-state">No journal data to analyze</p></div>
                    <div style="padding:0 14px 8px;"><span id="cfg-fb-backtest" style="font-size:11px;"></span></div>
                </div>
            </div>
        </div>
        <div class="reports-right">
            <div class="tab-bar">
                <button class="tab-btn active" onclick="switchTab('log-tab')">Live Log</button>
                <button class="tab-btn" onclick="switchTab('journal-tab')">Trade Journal</button>
                <button class="tab-btn" onclick="switchTab('analyzer-tab')">🎯 Negation Analyzer</button>
            </div>
            <div id="log-tab" class="tab-content active">
                <div class="section-panel">
                    <div class="section-header"><span>Live Log</span>
                        <div style="display:flex;align-items:center;gap:6px;">
                            <select onchange="setFilter('log',this.value)" class="filter-select">
                                <option value="all">All</option>
                                <option value="index">Index</option>
                                <option value="nifty50">Nifty 50</option>
                            </select>
                            <button class="clear-logs-btn" onclick="clearLogs()">Clear Logs</button>
                        </div>
                    </div>
                    <div class="log-box" id="log-body"></div>
                </div>
            </div>
            <div id="journal-tab" class="tab-content">
                <div class="section-panel">
                    <div class="section-header"><span>Trade Journal</span>
                        <div style="display:flex;align-items:center;gap:6px">
                            <select onchange="setFilter('journal',this.value)" class="filter-select">
                                <option value="all">All</option>
                                <option value="index">Index</option>
                                <option value="nifty50">Nifty 50</option>
                            </select>
                            <button class="clear-logs-btn" onclick="clearJournal()">Clear</button>
                        </div>
                    </div>
                    <div id="journal-body"><p class="empty-state">No journal entries yet</p></div>
                </div>
            </div>
            <div id="analyzer-tab" class="tab-content">
                <div class="section-panel">
                    <div class="section-header"><span>🎯 Negation Theory Interactive Analyzer</span></div>
                    <div style="padding:16px;">
                        <p style="font-size:12px;color:#8b949e;margin-bottom:16px;">
                            Enter any contract or spot symbol with your entry parameters to calculate exact Negation Theory Targets (T1/T2/T3) and 10% Max Loss SL instantly.
                        </p>
                        <div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(180px, 1fr));gap:12px;margin-bottom:16px;">
                            <div>
                                <label style="font-size:11px;color:#8b949e;display:block;margin-bottom:4px;">Symbol / Contract</label>
                                <input type="text" id="an-symbol" placeholder="e.g. WIPRO26AUG200CE or VEDL" style="width:100%;padding:8px;background:#0d1117;border:1px solid #30363d;color:#c9d1d9;border-radius:6px;font-size:12px;">
                            </div>
                            <div>
                                <label style="font-size:11px;color:#8b949e;display:block;margin-bottom:4px;">Entry Price (₹)</label>
                                <input type="number" step="any" id="an-entry" placeholder="e.g. 1.60 or 12.00" style="width:100%;padding:8px;background:#0d1117;border:1px solid #30363d;color:#c9d1d9;border-radius:6px;font-size:12px;">
                            </div>
                            <div>
                                <label style="font-size:11px;color:#8b949e;display:block;margin-bottom:4px;">Timeframe</label>
                                <select id="an-tf" style="width:100%;padding:8px;background:#0d1117;border:1px solid #30363d;color:#c9d1d9;border-radius:6px;font-size:12px;">
                                    <option value="30minute" selected>30min (Same Entry & Anchor TF)</option>
                                    <option value="75min">75min (Anchor TF)</option>
                                    <option value="60minute">60min (Anchor TF)</option>
                                    <option value="15minute">15min (Entry TF)</option>
                                    <option value="day">Daily</option>
                                </select>
                            </div>
                            <div>
                                <label style="font-size:11px;color:#8b949e;display:block;margin-bottom:4px;">Engine / Market</label>
                                <select id="an-engine" style="width:100%;padding:8px;background:#0d1117;border:1px solid #30363d;color:#c9d1d9;border-radius:6px;font-size:12px;">
                                    <option value="nifty50">Nifty 50 Stock Options</option>
                                    <option value="index">Index Options</option>
                                </select>
                            </div>
                        </div>
                        <button id="an-submit-btn" onclick="runNegationAnalysis()" style="background:#2ea043;color:#ffffff;border:none;padding:9px 18px;border-radius:6px;font-weight:600;cursor:pointer;font-size:13px;width:100%;">🔍 Analyze Negation Targets & SL</button>
                        
                        <div id="analyzer-results"></div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="last-updated" id="last-updated">Loading...</div>
</body>
</html>
"""

# ──────────────────────────────────────────────
#  FLASK ROUTES — API Endpoints
# ──────────────────────────────────────────────

@app.route("/")
def dashboard():
    return render_template_string(HTML_TEMPLATE, refresh=REFRESH_SECONDS, programs=PROGRAMS)

@app.route("/api/status")
def api_status():
    with data_lock:
        prog_status = {}
        for pid in PROGRAMS:
            pid_running = get_pid_for_program(pid) is not None
            prog_status[pid] = {
                "running": pid_running,
                "scans": cached_data["scans"].get(pid, []),
                "log_tail": cached_data["log_tail"].get(pid, []),
                "scan_summary": cached_data["scan_summary"].get(pid, {"anchors": {}, "abc_matches": {}})
            }
        cfg = load_config()
        return jsonify({
            "programs": prog_status,
            "positions": cached_data["positions"],
            "all_trades": cached_data["all_trades"],
            "active_positions": cached_data["active_positions"],
            "ltp": {str(k): v for k, v in cached_data["ltp"].items()},
            "journal": cached_data["journal"],
            "stats": cached_data["stats"],
            "config": cfg,
            "scan_display": cached_data["scan_display"],
            "live_execution": cached_data["live_execution"],
            "live_execution_index": cached_data["live_execution_index"]
        })

@app.route("/api/token/check")
def api_token_check():
    return jsonify(check_token_valid())

@app.route("/api/token/url")
def api_token_url():
    return jsonify({"url": get_login_url()})

@app.route("/api/token/exchange", methods=["POST"])
def api_token_exchange():
    data = request.get_json(force=True, silent=True)
    if not data or not data.get("request_token"):
        return jsonify({"ok": False, "error": "No request_token provided"})
    result = exchange_request_token(data["request_token"].strip())
    return jsonify(result)

@app.route("/api/backtest/mode", methods=["GET", "POST"])
def api_backtest_mode():
    if request.method == "POST":
        data = request.get_json(force=True, silent=True)
        enabled = bool(data.get("enabled", False))
        set_backtest_mode(enabled)
    return jsonify({"enabled": get_backtest_mode()})

@app.route("/api/config/<prog_id>", methods=["POST"])
def api_save_config(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({"ok": False, "error": "Invalid JSON"})
    save_config(prog_id, data)
    return jsonify({"ok": True})

@app.route("/api/scan/clear", methods=["POST"])
def api_scan_clear():
    now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    empty_scan = {"date": dt.now().strftime("%Y-%m-%d"),
                  "timestamp": now_str,
                  "cleared_at": now_str,
                  "staged_trades": [], "carry_forward": [], "active_live": []}
    for f in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
        try:
            with open(f, "w") as fh:
                json.dump(empty_scan, fh)
        except Exception:
            pass
    return jsonify({"ok": True})

def _format_pattern_result(p):
    if not p: return '-'
    p_str = str(p)
    if 'Engulf' in p_str:
        return 'BEAR_ENG' if 'Bear' in p_str or 'BEAR' in p_str else 'BULL_ENG'
    elif 'Two_Higher' in p_str or 'Higher_Highs' in p_str:
        return 'BULL_2HH'
    elif 'Two_Lower' in p_str or 'Lower_Lows' in p_str:
        return 'BEAR_2LL'
    elif 'HH_Sweep' in p_str or 'HH_sweep' in p_str:
        return 'BEAR_HH'
    elif 'Sweep' in p_str or 'LL' in p_str:
        return 'BULL_LL'
    elif 'Star' in p_str or 'Shooting' in p_str:
        return 'BEAR_STAR'
    elif 'Baby' in p_str or 'Hammer' in p_str:
        return 'BULL_HAM'
    elif 'Harami' in p_str:
        return 'BEAR_HAR' if 'Bear' in p_str or 'BEAR' in p_str else 'BULL_HAR'
    elif 'Base' in p_str:
        return 'BULL_BASE'
    elif p_str == 'SCAN_READY':
        return 'BULL_ENG'
    return p_str

def _format_timestamp(ts):
    if not ts: return '-'
    try:
        s = str(ts).split('+')[0].replace('T', ' ')
        p = s.split(' ')
        dp = p[0].split('-') if p[0] else []
        tp = p[1].split(':') if len(p) > 1 and p[1] else []
        if len(dp) == 3 and len(tp) >= 2:
            return f"{dp[2]}-{dp[1]}-{dp[0][-2:]} {tp[0]}:{tp[1]}"
        return s
    except Exception:
        return str(ts)

def _format_float(val, dec=2):
    if val is None or val == '' or val == '-':
        return '-'
    try:
        return f"{float(val):.{dec}f}"
    except Exception:
        return str(val)

@app.route("/api/scan/export", methods=["POST"])
def api_scan_export():
    try:
        import io
        from spot_enricher import extract_underlying_symbol, evaluate_spot_trend_and_t1
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Symbol", "Contract", "Side", "Entry", "SL", "T1", "T2", "T3",
                         "AncherT", "EntryTime", "Result", "CF", "RR", "Engine", "Status",
                         "Spot_Trend", "Spot_T1_Target"])
        files = [("Nifty 50", SCAN_DISPLAY_FILE), ("Index", SCAN_DISPLAY_INDEX_FILE)]
        spot_eval_cache = {}
        for label, path in files:
            full = path if os.path.isabs(path) else os.path.join(BASE_DIR, path)
            if not os.path.exists(full):
                continue
            with open(full) as f:
                data = json.load(f)
            for section_name, status_tag in [("staged_trades", "Staged"), ("active_live", "Active"), ("carry_forward", "CarryFwd")]:
                for t in data.get(section_name, []):
                    writer.writerow([
                        t.get("symbol", ""),
                        t.get("contract", ""),
                        t.get("side", ""),
                        _format_float(t.get("entry_spot")),
                        _format_float(t.get("current_sl")),
                        _format_float(t.get("t1")),
                        _format_float(t.get("t2")),
                        _format_float(t.get("t3")),
                        _format_timestamp(t.get("candle_a_time")),
                        _format_timestamp(t.get("entry_time")),
                        _format_pattern_result(t.get("pattern") or t.get("result")),
                        "Yes" if t.get("carry_forward") else "No",
                        _format_float(t.get("rr")),
                        label,
                        status_tag
                    ])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(csv_bytes, mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=scan_export_{dt.now().strftime('%d_%m_%y_%H%M')}.csv"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/journal/clear", methods=["POST"])
def api_journal_clear():
    try:
        if os.path.exists(JOURNAL_FILE):
            open(JOURNAL_FILE, "w").close()
    except Exception:
        pass
    return jsonify({"ok": True})

@app.route("/api/programs/<prog_id>/start", methods=["POST"])
def api_start(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    token = check_token_valid()
    if not token["valid"]:
        return jsonify({"ok": False, "error": token["reason"]})
    ok = start_program(prog_id)
    return jsonify({"ok": ok, "error": None if ok else "Start failed"})

@app.route("/api/programs/<prog_id>/stop", methods=["POST"])
def api_stop(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    ok = stop_program(prog_id)
    return jsonify({"ok": ok})

ANCHOR_SCAN_REQUEST_FILE = os.path.join("output", "monitor", "anchor_scan_request.txt")
ANCHOR_SCAN_STOP_FILE = os.path.join("output", "monitor", "anchor_scan_stop.txt")

@app.route("/api/anchor/scan", methods=["POST"])
def api_anchor_scan():
    data = request.get_json(silent=True) or {}
    engine = data.get("engine", "index")
    try:
        with data_lock:
            cached_data["anchor_status"]["running"] = True
            cached_data["anchor_status"]["engine"] = engine
            cached_data["anchor_status"]["requested_at"] = time.time()
        if os.path.exists(ANCHOR_SCAN_STOP_FILE):
            os.remove(ANCHOR_SCAN_STOP_FILE)
        # Launch a dedicated --anchor-only subprocess. We intentionally do NOT
        # write ANCHOR_SCAN_REQUEST_FILE here: a running engine also polls that
        # file in its main loop, which would cause the anchor scan to run twice.
        script = PROGRAMS.get(engine, {}).get("file")
        if script:
            script_path = os.path.join(BASE_DIR, script)
            subprocess.Popen([sys.executable, script_path, "--anchor-only"],
                             cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/stop", methods=["POST"])
def api_anchor_stop():
    try:
        os.makedirs(os.path.dirname(ANCHOR_SCAN_STOP_FILE), exist_ok=True)
        with open(ANCHOR_SCAN_STOP_FILE, "w") as f:
            f.write("stop")
        with data_lock:
            cached_data["anchor_status"]["running"] = False
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/status")
def api_anchor_status():
    with data_lock:
        st = dict(cached_data["anchor_status"])
    if not st.get("running"):
        still_running = os.path.exists(ANCHOR_SCAN_REQUEST_FILE) and not os.path.exists(ANCHOR_SCAN_STOP_FILE)
        if still_running:
            st["running"] = True
            if not st.get("engine"):
                try:
                    with open(ANCHOR_SCAN_REQUEST_FILE) as f:
                        st["engine"] = f.read().strip()
                except Exception:
                    pass
    return jsonify(st)

@app.route("/api/logs/clear", methods=["POST"])
def api_logs_clear():
    log_files = [INDEX_LOG_FILE, NIFTY50_LOG_FILE, DAILY_LOG_FILE]
    for lf in log_files:
        try:
            if os.path.exists(lf):
                open(lf, "w").close()
        except Exception:
            pass
    return jsonify({"ok": True})

@app.route("/api/trades")
def api_trades():
    engine = request.args.get("engine")
    active_only = request.args.get("active", "false").lower() == "true"
    if active_only:
        return jsonify(trade_db.get_active_trades(engine))
    return jsonify(trade_db.get_all_trades(engine))

@app.route("/api/export/monthly", methods=["POST"])
def api_export_monthly():
    result = run_monthly_export()
    return jsonify({"ok": True, **result})

@app.route("/api/live-execution/nifty50", methods=["GET", "POST"])
def api_live_execution():
    if request.method == "POST":
        enabled = request.get_json(force=True, silent=True).get("enabled", False)
        flag_path = os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG)
        if enabled:
            with open(flag_path, "w") as f:
                f.write("1")
        else:
            if os.path.exists(flag_path):
                os.remove(flag_path)
        with data_lock:
            cached_data["live_execution"] = enabled
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG))})

@app.route("/api/live-execution/index", methods=["GET", "POST"])
def api_live_execution_index():
    if request.method == "POST":
        enabled = request.get_json(force=True, silent=True).get("enabled", False)
        flag_path = os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG_INDEX)
        if enabled:
            with open(flag_path, "w") as f:
                f.write("1")
        else:
            if os.path.exists(flag_path):
                os.remove(flag_path)
        with data_lock:
            cached_data["live_execution_index"] = enabled
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG_INDEX))})

SL_TARGET_OVERRIDES_FILE = os.path.join(BASE_DIR, "output", "monitor", "sl_target_overrides.json")

@app.route("/api/edit-lock", methods=["POST"])
def api_edit_lock():
    data = request.json or {}
    sym = data.get("symbol")
    active = data.get("active", False)
    if sym:
        clean_s = str(sym).replace(" ", "").upper()
        if active:
            ACTIVE_EDIT_LOCKS.add(clean_s)
            logging.info(f"[EDIT LOCK ON] Automated exit execution paused for {clean_s}")
        else:
            ACTIVE_EDIT_LOCKS.discard(clean_s)
            logging.info(f"[EDIT LOCK OFF] Automated exit execution resumed for {clean_s}")
    return jsonify({"ok": True})

@app.route("/api/update-position", methods=["POST"])
def api_update_position():
    data = request.get_json(force=True, silent=True) or {}
    engine = data.get("engine", "nifty50")
    symbol = data.get("symbol", "")
    current_sl = data.get("current_sl")
    t1 = data.get("t1")
    t2 = data.get("t2")
    t3 = data.get("t3")
    if not symbol or (current_sl is None and t1 is None and t2 is None and t3 is None):
        return jsonify({"ok": False, "error": "symbol and at least one level required"}), 400
    vals = {}
    if current_sl is not None and str(current_sl).strip() != "": vals["current_sl"] = float(current_sl)
    if t1 is not None and str(t1).strip() != "": vals["t1"] = float(t1)
    if t2 is not None and str(t2).strip() != "": vals["t2"] = float(t2)
    if t3 is not None and str(t3).strip() != "": vals["t3"] = float(t3)
    vals["user_edited"] = True
    
    clean_target = str(symbol).replace(" ", "").upper()

    overrides = {}
    try:
        if os.path.exists(SL_TARGET_OVERRIDES_FILE):
            with open(SL_TARGET_OVERRIDES_FILE) as f:
                overrides = json.load(f)
    except Exception:
        overrides = {}

    for eng_k in (engine, "nifty50", "index"):
        overrides.setdefault(eng_k, {})[clean_target] = vals

    os.makedirs(os.path.dirname(SL_TARGET_OVERRIDES_FILE), exist_ok=True)
    with open(SL_TARGET_OVERRIDES_FILE, "w") as f:
        json.dump(overrides, f, indent=2)

    clear_executed_exit(symbol)
    clear_executed_exit(clean_target)
    ACTIVE_EDIT_LOCKS.discard(clean_target)

    matched = False
    with data_lock:
        update_keys = list(vals.keys())

        # 1. Update in-memory all_trades
        for t in cached_data.get("all_trades", []):
            t_sym = str(t.get("symbol") or "").replace(" ", "").upper()
            t_cnt = str(t.get("contract") or "").replace(" ", "").upper()
            if clean_target in (t_sym, t_cnt) or t_sym in clean_target or t_cnt in clean_target:
                matched = True
                for k in update_keys: t[k] = vals[k]
                tid = t.get("id")
                if tid:
                    trade_db.update_trade(tid, vals)

        # 2. Update in-memory positions
        for pos_key, pos in (cached_data.get("positions", {}).items() if isinstance(cached_data.get("positions"), dict) else enumerate(cached_data.get("positions", []))):
            if isinstance(pos, dict):
                p_sym = str(pos.get("symbol") or "").replace(" ", "").upper()
                p_cnt = str(pos.get("contract") or "").replace(" ", "").upper()
                if clean_target in (p_sym, p_cnt) or p_sym in clean_target or p_cnt in clean_target:
                    matched = True
                    for k in update_keys: pos[k] = vals[k]
                    tid = pos.get("id")
                    if tid:
                        trade_db.update_trade(tid, vals)

        # 3. Update in-memory active_positions so UI refreshes immediately
        for kp in cached_data.get("active_positions", []):
            k_sym = str(kp.get("symbol") or "").replace(" ", "").upper()
            k_cnt = str(kp.get("contract") or "").replace(" ", "").upper()
            if clean_target in (k_sym, k_cnt) or k_sym in clean_target or k_cnt in clean_target:
                for k in update_keys: kp[k] = vals[k]

        if not matched:
            contract = symbol
            exchange = "NSE"
            for kp in cached_data.get("active_positions", []):
                k_sym = str(kp.get("symbol") or "").replace(" ", "").upper()
                k_cnt = str(kp.get("contract") or "").replace(" ", "").upper()
                if clean_target in (k_sym, k_cnt) or k_sym in clean_target or k_cnt in clean_target:
                    contract = kp.get("contract", symbol)
                    exchange = kp.get("exchange", "NSE")
                    break
            is_stock = exchange == "NSE"
            trade_data = {"contract": contract, "entry_spot": 0, "position_type": "stock" if is_stock else "option"}
            trade_data.update(vals)
            tid = trade_db.create_trade(engine, symbol, trade_data)
            entry = {"symbol": symbol, "contract": contract, "id": tid, "engine": engine, "status": "ACTIVE", "position_type": "stock" if is_stock else "option"}
            entry.update(vals)
            cached_data["all_trades"].append(entry)
            cached_data["positions"][symbol] = entry
            logging.info(f"[OVERRIDE] Created new DB trade for {engine}/{symbol}")

        # Synchronize scan_display in memory and on disk so 1s polling preserves edit immediately
        disp_file = SCAN_DISPLAY_FILE if engine == "nifty50" else SCAN_DISPLAY_INDEX_FILE
        eng_disp = cached_data.get("scan_display", {}).get(engine, {})
        clean_sym = str(symbol).replace(" ", "").upper()
        if isinstance(eng_disp, dict):
            for cat in ["staged_trades", "active_live", "carry_forward"]:
                for item in eng_disp.get(cat, []):
                    if isinstance(item, dict):
                        i_sym = str(item.get("symbol") or "").replace(" ", "").upper()
                        i_cnt = str(item.get("contract") or "").replace(" ", "").upper()
                        if clean_sym in (i_sym, i_cnt) or i_sym in clean_sym or i_cnt in clean_sym:
                            for k in update_keys:
                                item[k] = vals[k]
                            if "current_sl" in vals and "entry_spot" in item and item.get("entry_spot"):
                                item["rr"] = round(calc_rr(item.get("entry_spot"), vals["current_sl"], vals.get("t1", item.get("t1")), vals.get("t2", item.get("t2"))), 2)
            if os.path.exists(disp_file):
                try:
                    with open(disp_file, "w") as fh:
                        json.dump(eng_disp, fh, indent=2)
                except Exception as fe:
                    logging.warning(f"Failed to update scan display file: {fe}")

    logging.info(f"Position override queued: {engine}/{symbol} {vals}")
    return jsonify({"ok": True})

# ──────────────────────────────────────────────
#  1-CLICK BUY SCANNED TRADE API
# ──────────────────────────────────────────────
@app.route("/api/buy-scanned-trade", methods=["POST"])
def api_buy_scanned_trade():
    try:
        data = request.json or {}
        symbol = data.get("symbol")
        contract = data.get("contract") or symbol
        side = data.get("side", "CE")
        entry_spot = float(data.get("entry_spot") or 0)
        current_sl = float(data.get("current_sl") or data.get("sl") or 0)
        t1 = float(data.get("t1") or 0)
        t2 = float(data.get("t2") or 0)
        t3 = float(data.get("t3") or 0)
        engine = data.get("engine", "nifty50")
        
        if not symbol:
            return jsonify({"ok": False, "error": "symbol is required"}), 400

        c_str = str(contract).upper()
        if "SENSEX" in c_str or "BSE" in c_str:
            exch = "BFO"
        elif "CE" in c_str or "PE" in c_str or "NIFTY" in c_str or "BANK" in c_str:
            exch = "NFO"
        else:
            exch = "NSE"

        logging.info(f"[1-CLICK BUY] Staging open-source trade setup for {contract}")

        trade_data = {
            "contract": contract,
            "entry_spot": entry_spot,
            "current_sl": current_sl,
            "t1": t1,
            "t2": t2,
            "t3": t3,
            "side": side,
            "pattern": "1CLICK_BUY",
            "position_type": "stock" if exch == "NSE" else "option",
            "user_edited": True
        }
        tid = trade_db.create_trade(engine, symbol, trade_data)
        clear_executed_exit(contract)

        return jsonify({
            "ok": True,
            "message": f"Successfully placed 1-Click BUY for {contract}" + (f" (Order ID: {order_id})" if order_id else ""),
            "trade_id": tid
        })
    except Exception as e:
        logging.error(f"1-Click Buy API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────
#  INTERACTIVE NEGATION ANALYZER API
# ──────────────────────────────────────────────
@app.route("/api/analyze-trade", methods=["POST"])
def api_analyze_trade():
    try:
        data = request.json or {}
        symbol = str(data.get("symbol", "")).strip().upper()
        entry_price = float(data.get("entry_price", 0)) if data.get("entry_price") else 0.0
        timeframe = str(data.get("timeframe", "75min")).strip()
        engine = str(data.get("engine", "nifty50")).strip()
        
        if not symbol:
            return jsonify({"ok": False, "error": "Valid Symbol or Contract Name required"}), 400

        analysis = derive_sl_targets_for_contract(None, symbol, entry_price, timeframe_entry, timeframe_anchor)
        if not analysis:
            sl_val = round(entry_price * 0.90, 2) if entry_price > 0 else 0.0
            analysis = {
                "entry_price": entry_price,
                "current_sl": sl_val,
                "t1": None, "t2": None, "t3": None,
                "pattern": "NEGATION_ESTIMATED"
            }

        resolved_entry = float(analysis.get("entry_price") or entry_price or 0.0)
        sl_val = analysis.get("current_sl", round(resolved_entry * 0.90, 2) if resolved_entry > 0 else 0.0)
        t1_val = analysis.get("t1")
        t2_val = analysis.get("t2")
        t3_val = analysis.get("t3")
        
        risk = (resolved_entry - sl_val) if (resolved_entry > 0 and sl_val < resolved_entry) else 0
        rr = round((t1_val - resolved_entry) / risk, 2) if (t1_val and risk > 0) else 0.0

        return jsonify({
            "ok": True,
            "symbol": symbol,
            "contract": symbol,
            "entry_price": resolved_entry,
            "current_sl": sl_val,
            "t1": t1_val if t1_val else "N/A",
            "t2": t2_val if t2_val else "N/A",
            "t3": t3_val if t3_val else "N/A",
            "rr": rr,
            "pattern": analysis.get("pattern", "NEGATION_DERIVED"),
            "engine": engine
        })
    except Exception as e:
        logging.error(f"Analyze Trade API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────
#  DAILY SELF-LEARNING TRADE JOURNAL API
# ──────────────────────────────────────────────
@app.route("/api/journal/get", methods=["GET"])
def api_journal_get():
    try:
        from daily_trade_journal import load_journal_entries
        return jsonify(load_journal_entries())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/sync", methods=["POST"])
def api_journal_sync():
    try:
        from daily_trade_journal import generate_daily_journal
        req = request.json or {}
        dt_str = req.get("date")
        entries = generate_daily_journal(dt_str)
        return jsonify({"ok": True, "count": len(entries), "entries": entries})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/update", methods=["POST"])
def api_journal_update():
    try:
        from daily_trade_journal import load_journal_entries, save_journal_entries
        data = request.json or {}
        symbol = data.get("symbol")
        date_str = data.get("date")
        remarks = data.get("remarks")
        lesson = data.get("lesson")
        if not symbol or not date_str:
            return jsonify({"ok": False, "error": "symbol and date required"}), 400
        entries = load_journal_entries()
        updated = False
        for e in entries:
            if e.get("Date") == date_str and (e.get("Symbol") == symbol or symbol in e.get("Symbol", "")):
                if remarks is not None: e["Analysis_Remarks"] = remarks
                if lesson is not None: e["Self_Learning_Lesson"] = lesson
                updated = True
        if updated:
            save_journal_entries(entries)
            return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "entry not found"}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

EXPORT_STATE_FILE = os.path.join(BASE_DIR, "output", "monitor", "export_state.json")

# ──────────────────────────────────────────────
#  MONTHLY EXPORT (trades to Excel archive)
# ──────────────────────────────────────────────

def run_monthly_export():
    import openpyxl
    from collections import defaultdict
    completed = trade_db.get_completed_trades()
    if not completed:
        return {"exported": 0, "sheets": []}
    groups = defaultdict(list)
    for t in completed:
        ts = t.get("exit_time") or t.get("updated_at") or t.get("created_at") or ""
        parts = ts.split(" ")[0].split("-") if " " in ts else ts.split("-")
        key = (parts[0], parts[1]) if len(parts) >= 2 else ("unknown", "00")
        groups[key].append(t)
    out_dir = os.path.join(BASE_DIR, "output", "exports")
    os.makedirs(out_dir, exist_ok=True)
    xl_path = os.path.join(out_dir, "trade_archive.xlsx")
    sheet_names = []
    if os.path.exists(xl_path):
        wb = openpyxl.load_workbook(xl_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    headers = ["ID", "Engine", "Symbol", "Pattern", "Status", "Entry Spot", "SL", "T1", "T2", "T3",
               "Trailing Stage", "Lot Size", "Position Size", "Entry Date", "Exit Date", "P&L %", "Side", "Contract"]
    for (year, month), trades in sorted(groups.items()):
        month_names = ["", "January","February","March","April","May","June","July","August","September","October","November","December"]
        sheet_name = f"{month_names[int(month)]} {year}" if month.isdigit() and 1 <= int(month) <= 12 else f"{month} {year}"
        sheet_names.append(sheet_name)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for t in trades:
            entry_dt = (t.get("created_at") or "").split(" ")[0]
            exit_dt = (t.get("exit_time") or t.get("updated_at") or "").split(" ")[0]
            ws.append([
                t.get("id", ""), t.get("engine", ""), t.get("symbol", ""), t.get("pattern", ""),
                t.get("status", ""), t.get("entry_spot", ""), t.get("current_sl", ""),
                t.get("t1", ""), t.get("t2", ""), t.get("t3", ""),
                t.get("trailing_stage", ""), t.get("lot_size", ""), t.get("position_size", ""),
                entry_dt, exit_dt,
                t.get("pnl_percent", ""), t.get("side", ""), t.get("contract", "")
            ])
    wb.save(xl_path)
    exported_ids = [t["id"] for t in completed]
    trade_db.remove_trades(exported_ids)
    return {"exported": len(completed), "sheets": sheet_names}

def auto_export_if_new_month():
    now = dt.now()
    current_month = now.strftime("%Y-%m")
    try:
        if os.path.exists(EXPORT_STATE_FILE):
            with open(EXPORT_STATE_FILE) as f:
                state = json.load(f)
                last = state.get("last_export_month", "")
        else:
            last = ""
        if current_month > last:
            result = run_monthly_export()
            if result["exported"] > 0:
                print(f"Auto-export: {result['exported']} trades to {', '.join(result['sheets'])}")
            with open(EXPORT_STATE_FILE, "w") as f:
                json.dump({"last_export_month": current_month}, f)
    except Exception as e:
        print(f"Auto-export error: {e}")

_last_eod_journal_triggered_date = None

def auto_eod_journal_scheduler():
    global _last_eod_journal_triggered_date
    while True:
        try:
            now = dt.now()
            today_str = now.strftime("%Y-%m-%d")
            # Trigger once daily at/after 15:35 IST on weekdays (Mon-Fri)
            if now.weekday() < 5 and (now.hour > 15 or (now.hour == 15 and now.minute >= 35)):
                if _last_eod_journal_triggered_date != today_str:
                    _last_eod_journal_triggered_date = today_str
                    logging.info(f"[AUTO EOD JOURNAL] Market closed. Auto-generating EOD trade journal for {today_str}...")
                    from daily_trade_journal import generate_daily_journal
                    generate_daily_journal(target_date=today_str)
                    logging.info(f"[AUTO EOD JOURNAL] Successfully completed EOD trade journal sync for {today_str}.")
        except Exception as e:
            logging.warning(f"[AUTO EOD JOURNAL] Error in scheduler: {e}")
        time.sleep(60)

def main():
    os.makedirs(os.path.join(BASE_DIR, "input"), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, "output", "logs"), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, "output", "monitor"), exist_ok=True)
    os.makedirs("logs", exist_ok=True)
    auto_export_if_new_month()
    worker = threading.Thread(target=refresh_data, daemon=True)
    worker.start()
    eod_worker = threading.Thread(target=auto_eod_journal_scheduler, daemon=True)
    eod_worker.start()
    print(f"Trading Control Center starting on http://localhost:{DASHBOARD_PORT}")
    print(f"Refresh interval: {REFRESH_SECONDS}s")
    print("Available programs:")
    for pid, p in PROGRAMS.items():
        print(f"  [{pid}] {p['name']}")
    app.run(host="0.0.0.0", port=DASHBOARD_PORT, debug=False, use_reloader=False)

if __name__ == "__main__":
    main()
